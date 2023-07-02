import logging
from math import ceil
from tkinter import ttk
from requests import get
import http.client
import json
import mimetypes
from tkinter import *
from glob import glob
from tkinter.messagebox import askyesno, showerror, showwarning, showinfo
from tkinter.filedialog import askopenfilename
import ast
import socket
import struct
import uuid
import random
import constant
from openpyxl import load_workbook
from account import Account
from account_db import Account_DB
from angel_db import AngelInstruments
from orders_db import Orders
from zerodha_db import ZerodhaInstruments
from risk_profile_db import RiskProfile
from datetime import datetime
from progress_bar import ProgressBarPanel
import csv
import queue
import copy

# from utility_treads import InternetUtility

Width, Height = 650, 460
imageTypes = [('Gif files', '.gif'),  # for file open dialog
              ('Ppm files', '.ppm'),  # plus jpg with a Tk patch,
              ('Pgm files', '.pgm'),  # plus bitmaps with BitmapImage
              ('All files', '*')]
sheetTypes = [('XLSX files', '.xlsx')]
jsonTypes = [('Json files', '.json')]

APIKEY = '1B55TGBb'
SECRETKEY = 'fee2a674-419e-46a2-9afc-25ee054ea54f'
CLIENT_ID = 'A1295683'
PASSWORD = 'Platinum@10'
CLIENTLOCALIP = '192.168.1.17'
CLIENTPUBLICIP = '122.172.83.83'
MACADDRESS = '16:58:63:8c:33:ce'
logging.basicConfig(level=logging.INFO,
                    format='%(process)d-%(levelname)s-%(message)s')

class CopyTraderGUI(Frame):
    def __init__(self, parent=NONE, msecs=3000, **args):
        Frame.__init__(self, parent, args)
        self.parent = parent
        self.parent.attributes('-fullscreen', True)
        self.screen_height = parent.winfo_screenheight() - 20
        self.screen_width = parent.winfo_screenwidth() - 200
        self.parent.attributes('-fullscreen', False)
        self.parent.protocol("WM_DELETE_WINDOW", lambda: (self.onQuit()))
        self.style = ttk.Style(self.parent)
        self.style.theme_use(self.style.theme_names()[0])
        self.selectedInstrumentData = ''
        self.instruments = []
        self.listOfAccounts = []
        self.listOfXLSXAccounts = []
        self.threaded_queue = queue.Queue()
        self.Orderframe = None
        self.parent.after(300, self.listen_for_result)
        self.parent.after(90000, self.check_for_auth_status)

        
        self.account_risk_vars = ['Low', 'Medium', 'High']
        self.is_place_order_panel_initial_load = True
        # self.parent.after(2000, self.simulate_result)
        self.is_view_order_win_initial_load = True
        self.is_order_modification_win_initial_load = True
        self.is_singleorder_exit_win_initial_load = True
        self.start_progress_bar()
        # if self.check_internet_basic():
        if True:
            self.load_accounts_from_db(True)
            self.makeWidgets()
            self.createListOfAccountsWidget()
            self.setup_database()
            self.setup_default_risk_profiles()
            self.place_order()
            self.pack(expand=YES, fill=BOTH)
        else:
            try:
                if showerror('Copy Trader', 'Please check if your device is connected to the internet and restart the application') == 'ok':
                    print('west')
                    self.pack(expand=YES, fill=BOTH)
                    self.update()
                    self.quit()
                    self.parent.destroy()
            except Exception as e:
                print(e)
        self.stop_progress_bar()
        self.queue = queue
        self.msecs = msecs
        self.beep = 1
        self.drawn = None

    def check_internet_basic(self, host="8.8.8.8", port=53, timeout=3):
        """ Host: 8.8.8.8 (google-public-dns-a.google.com)
        OpenPort: 53/tcp
        Service: domain (DNS/TCP) """

        try:
            socket.setdefaulttimeout(timeout)
            socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect(
                (host, port))
            return True
        except socket.error as ex:
            print(ex)
            return False
        except Exception as e:
            print(e)
            return False

    def makeWidgets(self):

        self.canvas = Frame(self, height=self.screen_height,
                            width=self.screen_width/2)
        self.canvas.pack_propagate(0)
        self.canvas.pack(side=LEFT, fill=BOTH, expand=YES, padx=5)
        self.menu = Menu(self)
        account = Menu(self.menu)
        account.add_command(label='Add Account',
                            command=self.addAccountScreen)
        account.add_command(label='Add Risk Rules',
                            command=self.add_risk_rules)
        self.menu.add_cascade(label='Account', menu=account)
        order = Menu(self.menu)
        # order.add_command(label='Place Order', command=self.place_order)
        # order.add_command(label='View Order', command=self.loadOrderScreen)
        order.add_command(label='View Positions',
                          command=self.loadPositionScreen)
        self.menu.add_cascade(label='Order', menu=order)
        self.menu.add_command(label='Quit', command=self.onQuit)
        self.master.configure(menu=self.menu)

    def fetch(self):
        print('Input => "%s"' + self.clientcode.get())  # get text

    def listen_for_result(self):
        """ Check if there is something in the queue. """

        try:
            # self.res
            task = self.threaded_queue.get(0)
            print(task, 'result generated', self.threaded_queue.qsize())
            is_task_done = False
            if type(task) is dict and task.get('type') and task['type'] == 'order' and task['data']:
                list_of_orders = self.order_db.get_orders()
                if len(list_of_orders) > 0:
                    for item in list_of_orders:
                        account_level_orders = json.loads(item[2])
                        for key in account_level_orders.keys():
                            print(
                                account_level_orders[key]['status'], key, '####')
                            if account_level_orders != None and account_level_orders[key]['status'] != 'error' and task['data']['account_id'] == key and task['data']['order_id'] == account_level_orders[key]['data']['order_id']:
                                if task['data']['meta'] != None and 'iceberg' in  task['data']['meta']:
                                    self.update_order_status(
                                        item[0], key, 'UPDATE','exchange_order_status')
                                    self.update_order_status(
                                        item[0], key, task['data']['meta']['iceberg']['total_quantity'] - task['data']['meta']['iceberg']['remaining_quantity'],'fill_quantity')
                                else:
                                    print(task['data']['account_id'], key, task['data']
                                        ['order_id'], account_level_orders[key]['data']['order_id'])
                                    self.update_order_status(
                                        item[0], key, task['data']['status'],'exchange_order_status')
                                    self.update_order_status(
                                        item[0], key, task['data']['filled_quantity'],'fill_quantity')
                                if task['data']['status'] == 'COMPLETE' or task['data']['status'] == 'UPDATE':
                                    self.recreate_open_position_tree_for_account(
                                        key)
                                is_task_done = True
                            if account_level_orders != None and account_level_orders[key]['status'] != 'error' and task['data']['account_id'] == key and task['data']['parent_order_id'] == account_level_orders[key]['data']['order_id']:
                                if task['data']['meta'] != None and 'iceberg' in  task['data']['meta']:
                                    status = 'UPDATE'
                                    filled_quant = task['data']['meta']['iceberg']['total_quantity'] - task['data']['meta']['iceberg']['remaining_quantity']
                                    if task['data']['status'] == 'REJECTED' or task['data']['status'] == 'CANCELLED':
                                        status = 'REJECTED'
                                    if task['data']['meta']['iceberg']['remaining_quantity'] == 0 and task['data']['status'] == 'COMPLETE':
                                        status = 'COMPLETE'
                                    if status != 'REJECTED' or status != 'CANCELLED':
                                        self.update_order_status(
                                            item[0], key, filled_quant,'fill_quantity')
                                    self.update_order_status(
                                        item[0], key, status,'exchange_order_status')


                                is_task_done = True
            if is_task_done == False:
                self.threaded_queue.put(task)

            self.parent.after(500, self.listen_for_result)
        except queue.Empty:
            print('empty queue')
            self.parent.after(1000, self.listen_for_result)

    def simulate_result(self):
        self.threaded_queue.put(
            # json.dumps(
                {"type":"order","id":"","data":{"account_id":"ZL3443","unfilled_quantity":0,"checksum":"","placed_by":"ZL3443","order_id":"230324202606793","exchange_order_id":"2100000024359273","parent_order_id":'null',"status":"UPDATE","status_message":'null',"status_message_raw":'null',"order_timestamp":"2023-03-24 15:24:36","exchange_update_timestamp":"2023-03-24 15:24:36","exchange_timestamp":"2023-03-24 10:43:29","variety":"regular","exchange":"NFO","tradingsymbol":"TCS23APR3360CE","instrument_token":38196482,"order_type":"LIMIT","transaction_type":"BUY","validity":"DAY","product":"NRML","quantity":175,"disclosed_quantity":0,"price":6.25,"trigger_price":0,"average_price":0,"filled_quantity":random.randint(30,100),"pending_quantity":175,"cancelled_quantity":175,"market_protection":0,"meta":{},"tag":'null',"guid":"66270XLptJ40MwQGlH"}}
                # )
                )
        self.parent.after(2000, self.simulate_result)
    def check_for_auth_status(self):
        # check_user_token
        try:
            print('auth check')
            is_account_logged_out = True
            for acc in self.listOfAccounts:
                if acc.get_auth_status() == 'Logged In':
                    # print(acc.check_user_token())
                    if acc.check_user_token() == 403:
                        acc.authStatus == 'Logged Out'
                        is_account_logged_out = True
                        showerror('Copy Trader','Account ' + acc.client_id + ' is no longer logged in, please login again to continue')
            self.parent.after(90000, self.check_for_auth_status)
            if is_account_logged_out:
                self.recreate_tree()
        except Exception as e:
            print(e)
    def loginTest(self):

        conn = http.client.HTTPSConnection(
            "apiconnect.angelbroking.com")
        # payload = "{\n\"clientcode\":\"CLIENT_ID\",\n\"password\":\"CLIENT_PASSWORD\"\n,\n\"totp\":\"TOTP_CODE\"\n}"
        payload = {
            'clientcode': CLIENT_ID,
            'password': PASSWORD,
            'totp': self.totp.get()
        }
        json_data = json.dumps(payload)
        """ s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        print('My ClientLocalIP:{}',s.getsockname()[0])
        ip = get('https://api.ipify.org').content.decode('utf8')
        print('My public IP address is: {}'.format(ip))
        print ("The formatted MAC address is : ", end="")
        print (':'.join(['{:02x}'.format((uuid.getnode() >> elements) & 0xff)
        for elements in range(0,2*6,2)][::-1])) """
        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'X-UserType': 'USER',
            'X-SourceID': 'WEB',
            # 'X-ClientLocalIP': s.getsockname()[0],
            'X-ClientLocalIP': CLIENTLOCALIP,
            # 'X-ClientPublicIP': format(ip),
            'X-ClientPublicIP': CLIENTLOCALIP,
            'X-MACAddress': MACADDRESS,
            'X-PrivateKey': APIKEY
        }
        # s.close()

        conn.request("POST", constant.LOGIN,
                     json_data,
                     headers)

        res = conn.getresponse()
        data = res.read()
        print(res.status, data.decode("utf-8"))

    


    def canvas_frame_reconfigurer(self, canvas):
        '''Reset the scroll region to encompass the inner frame'''
        canvas.configure(scrollregion=canvas.bbox("all"))      

    def place_order(self):

        self.call_progressbar()
        if self.is_place_order_panel_initial_load == False:
            self.place_order_var.destroy()

        if self.is_place_order_panel_initial_load == True:
            self.is_place_order_panel_initial_load = False
        self.orderObject = ''
        self.variety = StringVar()
        self.variety.set('NORMAL')
        self.transactiontype = StringVar()
        self.transactiontype.set('BUY')
        self.ordertype = StringVar()
        self.ordertype.set('LIMIT')
        self.producttype = StringVar()
        self.producttype.set('CARRYFORWARD')
        self.duration = StringVar()
        self.duration.set('DAY')
        self.exchange = StringVar()
        self.exchange.set('NFO')
        self.place_order_var = Frame(self,
                                     height=self.screen_height, width=900,
                                     #  padx=20,
                                     pady=15
                                     )
        # win.pack_propagate(0)
        # win.title('Order')
        self.place_order_var.pack(side=TOP)
        # win.config()
        canvas = Canvas(self.place_order_var, borderwidth=0,
                        background="#e0e0e0", width=210)
        accountSettingsFrame = LabelFrame(
            canvas, text='Selected Accounts', height=300, width=200, padx=2)
        accountSettingsFrame.pack(side=RIGHT, fill=Y)

        sbar = Scrollbar(self.place_order_var,
                         orient="vertical", command=(canvas.yview))
        canvas.configure(yscrollcommand=sbar.set)
        sbar.pack(side=RIGHT, fill=Y)
        canvas.pack(side=RIGHT, fill=Y)
        canvas.create_window((1, 1), window=accountSettingsFrame, anchor="nw")
        accountSettingsFrame.bind(
            "<Configure>", lambda event, canvas=canvas: self.canvas_frame_reconfigurer(canvas))

        account_orderplacement_panel = {}
        account_riskpanel = {}
        account_quantity_panel = {}
        account_risk_setting = {}
        account_moded_risk_amount = {}
        self.account_iceberg_lot = {}
        self.account_iceberg_quant = {}
        for acc in self.listOfAccounts:
            if acc.get_auth_status() == 'Logged In':
                account_frame = LabelFrame(
                    accountSettingsFrame, text=str(acc.client_id), padx=10)
                account_frame.pack(side=TOP, fill=X)
                non_ice_frame = Frame(account_frame)
                account_orderplacement_panel[acc.client_id] = BooleanVar()
                account_riskpanel[acc.client_id] = BooleanVar()
                account_quantity_panel[acc.client_id] = DoubleVar()
                self.account_iceberg_lot[acc.client_id] = IntVar()
                self.account_iceberg_quant[acc.client_id] = IntVar()
                account_risk_setting[acc.client_id] = acc.risk_setting
                account_orderplacement_panel[acc.client_id].set(True)
                account_riskpanel[acc.client_id].set(True)
                account_moded_risk_amount[acc.client_id] = acc.order_level_risks
                local_risk_checkbox = Checkbutton(
                    non_ice_frame, text='Risk', var=(account_riskpanel[acc.client_id]))
                local_risk_checkbox.pack(side=LEFT)
                # local_risk_checkbox.select()
                local_checkbox = Checkbutton(
                    non_ice_frame, text='Active', var=(account_orderplacement_panel[acc.client_id]))
                local_checkbox.pack(side=LEFT)
                
                local_entry = Entry(
                    non_ice_frame, textvariable=account_quantity_panel[acc.client_id], width=10)
                local_entry.pack(side=LEFT)
                local_entry.bind(
                    '<KeyRelease>', (lambda e: self.calculate_iceberg_lots(
                        account_quantity_panel
                    )))
                iceberg_frame = Frame(
                    account_frame
                )
                
                iceberg_leg_label = Label(
                    iceberg_frame, text='Leg', width=5)
                iceberg_leg_label.pack(side=LEFT, fill=NONE)
                iceberg_leg_entry = Entry(
                    iceberg_frame,
                    textvariable=self.account_iceberg_lot[acc.client_id],
                    width=4)
                iceberg_leg_entry.pack(side=LEFT, fill=NONE)
                
                iceberg_leg_entry.bind(
                    '<KeyRelease>', (lambda e: self.calculate_iceberg_lots(
                        account_quantity_panel
                    )))
                iceberg_quantity_label = Label(
                    iceberg_frame, text='Qty',
                    width=5)
                iceberg_quantity_label.pack(side=LEFT, fill=NONE)
                iceberg_quantity_entry = Entry(
                    iceberg_frame,
                    textvariable=self.account_iceberg_quant[acc.client_id],
                    width=6)
                iceberg_quantity_entry.pack(side=LEFT, fill=NONE)
                non_ice_frame.pack(side=TOP, fill=X)
                iceberg_frame.pack(side=LEFT, fill=X)

                # local_checkbox.select()
                print('>>>>>>>>>>>>>>')
                print(account_orderplacement_panel[acc.client_id],
                      account_orderplacement_panel[acc.client_id].get())
                print(account_riskpanel[acc.client_id],
                      account_riskpanel[acc.client_id].get())
                print('<<<<<<<<<<<<<<')
        Isearch = Frame(self.place_order_var, height=300, width=400, pady=2)
        labI = Label(Isearch, width=15, text='Search Instrument')
        self.searchInstrumentItem = Entry(Isearch, width=40)
        self.searchInstrumentItem.bind(
            '<KeyRelease>', (lambda e: self.searchInstruments()))
        Isearch.pack(side=TOP)
        labI.pack(side=LEFT)
        self.searchInstrumentItem.pack(side=LEFT,
                                       expand=YES)
        instrument = Frame(self.place_order_var, height=300, width=400)
        labI = Label(instrument, width=20, text='Instruments')
        sbar = Scrollbar(instrument)
        self.listBoxOfInstruments = Listbox(instrument,
                                            relief=SUNKEN, selectmode=SINGLE, exportselection=False)
        sbar.config(command=(self.listBoxOfInstruments.yview))
        self.listBoxOfInstruments.config(yscrollcommand=(sbar.set))
        sbar.pack(side=RIGHT, fill=Y)
        labI.pack(side=TOP, expand=None, fill=X)
        self.listBoxOfInstruments.pack(side=LEFT,
                                       expand=YES,
                                       fill=BOTH)
        instrument.pack(side=TOP, fill=X)
        db_instrument_list = self.angel_db.get_instruments_data()
        self.instrumentsData = db_instrument_list
        if len(db_instrument_list) != 0:
            for instrument in db_instrument_list:
                self.listBoxOfInstruments.insert(END, instrument[1])

        self.listBoxOfInstruments.bind(
            '<ButtonRelease-1>', (lambda e: self.selectInstrument(
                account_moded_risk_amount,
                account_quantity_panel,
                account_risk_setting
            )))
        selectedInstrument = Frame(
            self.place_order_var, height=300, width=400, pady=2)
        labP = Label(selectedInstrument, width=20,
                     text='Selected Instrument')
        self.selectedInstrument = Entry(selectedInstrument,
                                        width=40, text='Selected Instrument')
        selectedInstrument.pack(side=TOP)
        labP.pack(side=LEFT)
        self.selectedInstrument.pack(side=LEFT, expand=YES)

        self.order_level_risk_checkbox = BooleanVar()
        self.order_level_risk_checkbox.set(False)
        self.order_level_risk_category = StringVar()
        self.order_level_risk_category.set('low')
        risk_panel = LabelFrame(
            self.place_order_var, height=300, width=400, text='Risk management')
        risk_panel.pack(side=TOP, fill=X)
        Radiobutton(risk_panel, text='High', command=(lambda: self.multiplyLots(
            account_moded_risk_amount,
            account_quantity_panel,
            account_risk_setting
        )), variable=(
            self.order_level_risk_category), value='high').pack(side=LEFT)
        Radiobutton(risk_panel, text='Medium', command=(lambda: self.multiplyLots(
            account_moded_risk_amount,
            account_quantity_panel,
            account_risk_setting
        )), variable=(
            self.order_level_risk_category), value='medium').pack(side=LEFT)
        Radiobutton(risk_panel, text='Low', command=(lambda: self.multiplyLots(
            account_moded_risk_amount,
            account_quantity_panel,
            account_risk_setting
        )), variable=(
            self.order_level_risk_category), value='low').pack(side=LEFT)
        Checkbutton(risk_panel, text='Disable Risk Management',
                    command=(lambda: self.multiplyLots(
                        account_moded_risk_amount,
                        account_quantity_panel,
                        account_risk_setting
                    )), padx=30, variable=(self.order_level_risk_checkbox)).pack(side='left')
        combo_frame1 = Frame(self.place_order_var, pady=4)
        exchange = LabelFrame(combo_frame1, height=300,
                              width=400, text='Exchange')
        exchange.pack(side=LEFT)
        radio1 = Radiobutton(exchange, text='BSE EQ', command=(
            lambda: self.orderType()), variable=(self.exchange), value='BSE')
        radio1.pack(side=LEFT)
        radio2 = Radiobutton(exchange, text='NSE EQ', command=(
            lambda: self.orderType()), variable=(self.exchange), value='NSE')
        radio2.pack(side=LEFT)
        radio3 = Radiobutton(exchange, text='NFO', command=(
            lambda: self.orderType()), variable=(self.exchange), value='NFO')
        radio3.pack(side=LEFT)

        orderType = LabelFrame(
            combo_frame1, height=300, width=100, text='Order type')
        orderType.pack(side=LEFT)
        rad1 = Radiobutton(orderType, text='Market', command=(
            lambda: self.orderType()), variable=self.ordertype, value='MARKET')
        rad1.pack(side=LEFT)
        rad2 = Radiobutton(orderType, text='Limit', command=(
            lambda: self.orderType()), variable=self.ordertype, value='LIMIT')
        rad2.pack(side=LEFT)
        rad2 = Radiobutton(orderType, text='SL', command=(
            lambda: self.orderType()), variable=self.ordertype, value='SL')
        rad2.pack(side=LEFT)

        combo_frame1.pack(side=TOP, fill=X)

        combo_frame2 = Frame(self.place_order_var, pady=5)

        producttype = LabelFrame(combo_frame2,
                                 height=300, width=400, text='Product type')
        producttype.pack(side=LEFT)
        rad1 = Radiobutton(producttype, text='CNC', command=(
            lambda: self.orderType()), variable=(self.producttype), value='DELIVERY')
        rad1.pack(side=LEFT)
        rad2 = Radiobutton(producttype, text='NRML', command=(
            lambda: self.orderType()), variable=(self.producttype), value='CARRYFORWARD')
        rad2.pack(side=LEFT)
        rad3 = Radiobutton(producttype, text='MIS', command=(
            lambda: self.orderType()), variable=(self.producttype), value='INTRADAY')
        rad3.pack(side=LEFT)

        buySellFrame = LabelFrame(
            combo_frame2, height=300, width=100, text='Buy/Sell')
        buySellFrame.pack(side=LEFT)
        radio1 = Radiobutton(buySellFrame, text='Buy', command=(
            lambda: self.orderType()), variable=(self.transactiontype), value='BUY')
        radio1.pack(side=LEFT)
        radio2 = Radiobutton(buySellFrame, text='Sell', command=(
            lambda: self.orderType()), variable=(self.transactiontype), value='SELL')
        radio2.pack(side=LEFT)

        order_variety = LabelFrame(
            combo_frame2, height=300, width=100, text='Type')
        order_variety.pack(side=LEFT)
        radio6 = Radiobutton(order_variety, text='Regular', command=(
            lambda: self.orderType()), variable=(self.variety), value='NORMAL')
        radio6.pack(side=LEFT)
        radio7 = Radiobutton(order_variety, text='Iceberg', command=(
            lambda: self.orderType()), variable=(self.variety), value='ICEBERG')
        radio7.pack(side=LEFT)

        combo_frame2.pack(side=TOP, fill=X)

        iceberg_frame = Frame(self.place_order_var, height=300, pady=5)
        # iceberg_frame.pack(side=TOP, fill=X)

        labIL = Label(iceberg_frame, width=15, text='Iceberg Leg')
        self.entIL = Spinbox(iceberg_frame, from_=2, to=10, values=(2, 3, 4, 5, 6, 7, 8, 9, 10),
                             width=5, textvariable=IntVar(value=2), wrap=False, command=(),
                             state='readonly')
        labIL.pack(side=LEFT)
        self.entIL.pack(side=LEFT)

        labIQ = Label(iceberg_frame, width=15, text='Iceberg Quantity')
        self.entIQ = Entry(iceberg_frame, width=5, text='Iceberg Quantity')
        labIQ.pack(side=LEFT)
        self.entIQ.pack(side=LEFT)


        radio6.configure(command=(lambda: self.toggle_iceberg(iceberg_frame, 'disable', account_riskpanel,
                                                              account_quantity_panel,
                                                              account_risk_setting)))
        radio7.configure(command=(lambda: self.toggle_iceberg(iceberg_frame, 'normal', account_riskpanel,
                                                              account_quantity_panel,
                                                              account_risk_setting)))

        quantity = Frame(self.place_order_var, height=300,
                         width=100, padx=5, pady=5)
        labL = Label(quantity, width=10, text='Quantity')
        self.entL = Entry(quantity, width=5, text='Lot size')
        labL.pack(side=LEFT)
        self.entL.pack(side=LEFT)
        self.entL.configure(state='normal')
        self.entL.delete(0, END)
        self.entL.insert(0, int(0))
        self.entL.configure(state='disable')
        labM = Label(quantity, width=7, text='Lot')
        current_value_quant = IntVar(value=1)
        self.entM = Spinbox(quantity, from_=0, to=50, values=(0, 10, 20, 30, 40, 50),
                            width=5, textvariable=current_value_quant, wrap=False,
                            command=(lambda: self.multiplyLots(
                                account_moded_risk_amount,
                                account_quantity_panel,
                                account_risk_setting
                            )))
        self.entM.bind('<KeyRelease>', (lambda e: self.multiplyLots(
            account_moded_risk_amount,
            account_quantity_panel,
            account_risk_setting
        )))
        labM.pack(side=LEFT)
        self.entM.pack(side=LEFT)
        labQ = Label(quantity, width=15, text='Total Quantity')
        self.entQ = Entry(quantity, width=10, text='Total Quantity')
        labQ.pack(side=LEFT)
        self.entQ.pack(side=LEFT)

        quantity.pack(side=TOP, fill=X)

        price_combo1 = Frame(self.place_order_var,
                             height=300, width=100, padx=1, pady=5)
        labP = Label(price_combo1, width=5, text='Price')
        self.entP = Entry(price_combo1, width=10, text='Enter Price')
        labP.pack(side=LEFT)
        self.entP.pack(side=LEFT, padx=4)
        price_combo1.pack(side=TOP, fill=X)
        # self.toggle_iceberg(iceberg_frame, 'disable', account_riskpanel,
        #                     account_quantity_panel,
        #                     account_risk_setting)
        update_price = Button(price_combo1, text='Update Price', command=(
            lambda: self.update_last_traded_price()))
        # cancelBtn = Button(btnsFrame, text='Cancel', command=())
        update_price.pack(side=LEFT)

        labP = Label(price_combo1, width=10, text='Stop Loss')
        self.entSL = Entry(price_combo1, width=10, text='Enter Stop Loss')
        labP.pack(side=LEFT)
        self.entSL.pack(side=LEFT)
        price_combo1.pack(side=TOP, fill=X)

        btnsFrame = Frame(self.place_order_var, height=300,
                          width=400, padx=5, pady=4)
        confirmBtn = Button(btnsFrame, text='Place Order', command=(
            lambda: self.execute_order(
                account_orderplacement_panel,
                account_quantity_panel,
                account_riskpanel,
                self.place_order_var
            )))
        # cancelBtn = Button(btnsFrame, text='Cancel', command=())
        btnsFrame.pack(side=TOP, fill=X)
        confirmBtn.pack(side=LEFT, pady=4, padx=4)
        # cancelBtn.pack(side=LEFT, pady=4, padx=4)
        self.stop_progressbar()

    def toggle_iceberg(self, iceberg_frame, state, account_riskpanel, account_quantity_panel, account_risk_setting):
        if state == 'normal':
            self.ordertype.set('LIMIT')
            self.entIL.configure(state='readonly')
            self.entM.delete(0, END)
            self.entM.insert(0, int(5))

        else:
            self.entIL.configure(state=state)
            self.entM.delete(0, END)
            self.entM.insert(0, int(0))

        # self.ordertype.set('MARKET')  
        self.entIQ.configure(state=state)
        self.multiplyLots(account_riskpanel,
                          account_quantity_panel,
                          account_risk_setting)

    def multiplyLots(self, riskpanel, quantity_panel,rp):
        """  account_riskpanel,
            account_quantity_panel,
            account_risk_setting """
        if self.entM.get() != '' and self.entL.get() != '':
            print(type(self.entM.get()), type(self.entL.get()))
            self.entQ.delete(0, END)
            self.entQ.insert(0, int(float(self.entM.get()))
                             * int(float(self.entL.get())))
        print(self.entM)
       
        if quantity_panel.values() != 0:
            for elem in quantity_panel.keys():
                quant = 0
                if self.order_level_risk_checkbox.get() == False:
                    quant = round(float(0 if self.entM.get() == '' else self.entM.get())
                            * riskpanel[elem][self.order_level_risk_category.get().lower()] / 100)
                else:
                    quant = round(float(0 if self.entM.get() == '' else self.entM.get()))
                if self.variety.get() == 'ICEBERG':
                    if quant < 5:
                        quant = 5
                    self.account_iceberg_lot[elem].set(3)
                    self.account_iceberg_quant[elem].set(round((quant/3)) * int(self.entL.get()))
                else:
                
                    self.account_iceberg_lot[elem].set(0)
                    self.account_iceberg_quant[elem].set(0)
                quantity_panel[elem].set(quant)
                # print(
                #     elem,
                #     risk_setting[elem],
                #     self.entM.get(),
                #     account_risk_matrix[risk_setting[elem]],
                #     self.order_level_risk_category.get()
                # )
    def calculate_iceberg_lots(self,quantity_panel):
        if self.entM.get() != '' and self.entL.get() != '':
            print(type(self.entM.get()), type(self.entL.get()))
            self.entQ.delete(0, END)
            self.entQ.insert(0, int(float(self.entM.get()))
                             * int(float(self.entL.get())))
        print(self.entM)
       
        if quantity_panel.values() != 0:
            for elem in quantity_panel.keys():
                quant = int(quantity_panel[elem].get())
                if self.variety.get() == 'ICEBERG':
                    if quant < 5:
                        quant = 5
                    print(round((quant/self.account_iceberg_lot[elem].get())),quant,self.account_iceberg_lot[elem].get(),int(self.entL.get()))
                    self.account_iceberg_quant[elem].set(ceil((quant/self.account_iceberg_lot[elem].get())) * int(self.entL.get()))
                else:
                    self.account_iceberg_lot[elem].set(0)
                    self.account_iceberg_quant[elem].set(0)
                
        pass
    def loadPositionScreen(self):
        if self.is_view_order_win_initial_load == False:
            self.view_order_win.destroy()
            # self.view_order_win.update()
        if self.is_view_order_win_initial_load == True:
            self.is_view_order_win_initial_load = False

        self.view_order_win = Toplevel(self, height=self.screen_height - 20, width=self.screen_width - 20, padx=10,
                                       pady=10)
        self.view_order_win.pack_propagate(0)
        self.view_order_win.title('View Orders')
        frame_width = int(
            (self.screen_width)/2)
        self.positionscreen_button_orders = Frame(
            self.view_order_win, padx=10, pady=10)
        refresh_orders = Button(
            self.positionscreen_button_orders, text='Refresh Orders', command=self.refresh_orders)
        # refresh_orders.configure(state='disable')
        refresh_orders.pack(side=LEFT, fill=NONE)

        modify_orders = Button(
            self.positionscreen_button_orders, text='Modify Orders', command=self.modify_multiple)
        # modify_orders.configure(state='disable')
        modify_orders.pack(side=LEFT, fill=NONE)

        delete_orders = Button(
            self.positionscreen_button_orders, text='Delete Orders', command=self.delete_multiple)
        # modify_orders.configure(state='disable')
        delete_orders.pack(side=LEFT, fill=NONE)

        delete_orders = Button(
            self.positionscreen_button_orders, text='Delete Order History', command=self.delete_order_history)
        # modify_orders.configure(state='disable')
        delete_orders.pack(side=LEFT, fill=NONE)

        add_positions = Button(
            self.positionscreen_button_orders, text='Add Positions', command=lambda: self.excecute_instruments_multiple('add'))
        # add_positions.configure(state='disable')
        add_positions.pack(side=RIGHT, fill=NONE)

       

        exit_positions_instruments = Button(
            self.positionscreen_button_orders, text='Exit Instrument', command=self.excecute_instruments_multiple)
        # exit_positions.configure(state='disable')
        exit_positions_instruments.pack(side=RIGHT, fill=NONE)

        refresh_positions = Button(self.positionscreen_button_orders, text='Refresh Positions',
                                   command=self.recreate_open_position_tree)
        # refresh_positions.configure(state='disable')
        refresh_positions.pack(side=RIGHT, fill=NONE)

        self.positionscreen_button_orders.pack(side=TOP, fill=X)

        self.runningOrdersFrame = Frame(self.view_order_win,
                                        width=frame_width,
                                        borderwidth=1, relief=RIDGE)
        column_width = int(frame_width/5)
        self.runningOrdersTree = ttk.Treeview(self.runningOrdersFrame, column=('TIME', 'INSTRUMENT', 'TYPE', 'QUANTITY','FILL'), show='headings',
                                              selectmode='extended')
        self.runningOrdersTree.heading('TIME', text='Timestamp')
        self.runningOrdersTree.heading('INSTRUMENT', text='Instrument')
        self.runningOrdersTree.heading('TYPE', text='Type')
        self.runningOrdersTree.heading('QUANTITY', text='Quantity')

        pos = 1
        col_width = self.runningOrdersTree.winfo_width()
        
        col = 0
        for acc in self.runningOrdersTree['columns']:
            self.runningOrdersTree.column(
                acc, anchor=CENTER, width=column_width)
            col += 1
            pass

        list_of_orders_tuples = self.order_db.get_orders()
        for item in list_of_orders_tuples:
            loaded_order = json.loads(item[1])
            account_level_orders = json.loads(item[2])
            self.runningOrdersTree.insert('', 'end', text=pos, iid=item[0],
                                          values=(
                                              item[0],
                                              loaded_order['tradingsymbol'], loaded_order['transactiontype'],
                                              loaded_order['quantity']), open=True, tags=('order',))
            # order_level_label_frame = Frame(self.runningOrdersFrame,width=400,borderwidth=3,relief=RIDGE)
            # Label(order_level_label_frame,text=
            # item[0] + ' \t' + loaded_order['tradingsymbol'] + '  \t' + loaded_order['transactiontype'] + '  \t' + loaded_order['quantity']
            # ,pady=10,font=(None, 10),wraplength=int(self.screen_width / 2),anchor='w').pack(side=TOP,fill=X)
            pos += 1
            # print(item[0], 'Checker')
            for key in account_level_orders.keys():
                # fill_string = 
                self.runningOrdersTree.insert(item[0], 'end', iid=item[0] + '#' + key,
                                              values=(
                    key,
                    account_level_orders[key]['exchange_order_status'] if account_level_orders[key].get(
                        'exchange_order_status') else 'UNKOWN',
                    '{0}/{1}'.format(account_level_orders[key]['fill_quantity'],account_level_orders[key]['order_quantity']) if account_level_orders[key].get(
                        'fill_quantity') else 0,
                    # 'Unknown'
                    'Modify',
                    'Delete'), open=True, tags=('content',account_level_orders[key]['exchange_order_status'] if account_level_orders[key].get(
                        'exchange_order_status') else 'UNKOWN'))

        treeScroll = ttk.Scrollbar(self.runningOrdersFrame)
        treeScroll.configure(command=(self.runningOrdersTree.yview))
        self.runningOrdersTree.configure(yscrollcommand=(treeScroll.set))
        self.runningOrdersTree.tag_configure('order', background='#ecf2fe', font=(
            None, 11))
        # self.runningOrdersTree.tag_configure('order', background='#000000', font=(
        #     None, 11),foreground='#ffffff')
        self.runningOrdersTree.tag_configure('content', font=(
            None, 9))
        self.runningOrdersTree.tag_configure('COMPLETE', font=(
            None, 9),foreground='#33ab07')
        self.runningOrdersTree.tag_configure('REJECTED', font=(
            None, 9),foreground='#ab0722')
        self.runningOrdersTree.tag_configure('UPDATE', font=(
            None, 9),foreground='#07916a')
        self.runningOrdersTree.tag_configure('CANCELLED', font=(
            None, 9),foreground='#ab0722')
        self.runningOrdersTree.tag_configure('OPEN PENDING', font=(
            None, 9),foreground='#f58f14')

        self.runningOrdersTree.bind(
            '<ButtonRelease-1>', lambda event: self.open_positions_tree_click_event(event, 'orders'))
        self.runningOrdersTree.pack(side=LEFT, fill=BOTH, anchor='ne')
        treeScroll.pack(side=LEFT, fill=Y)
        self.runningOrdersFrame.pack(side=LEFT, fill=Y, anchor='ne')

        self.openPositionsFrame = Frame(self.view_order_win, width=int(
            (self.screen_width)/2), borderwidth=3, relief=RIDGE)
        # if self.openPositionsTree != None and len(self.openPositionsTree.get_children()) != 0:
        #     self.openPositionsTree.delete(
        #         *self.openPositionsTree.get_children())
        self.openPositionsTree = ttk.Treeview(self.view_order_win, column=('ACCOUNT_NO', 'PRODUCT', 'QUANTITY', 'INSTRUMENT', 'TYPE'), show='headings',
                                              selectmode='extended')
        self.openPositionsTree.heading('ACCOUNT_NO', text='Account No')
        self.openPositionsTree.heading('PRODUCT', text='Product')
        self.openPositionsTree.heading('QUANTITY', text='Quantity')

        for acc in self.openPositionsTree['columns']:
            self.openPositionsTree.column(
                acc, anchor=CENTER, width=int(frame_width/5))
            pass
        pol = 1
        for account in self.listOfAccounts:
            if account.authStatus == 'Logged In':
                positions = account.get_positions()

                self.openPositionsTree.insert('', 'end', iid=account.client_id, text=pol,
                                            values=(
                                                account.client_id), open=True, tags=('order',))
                if (type(positions) is dict) and positions.get('status') and positions['status'] == 'success':
                    for position in positions['data']['net']:
                        # if len(self.openPositionsTree.get_children(account.client_id)) == 0:
                            self.openPositionsTree.insert(account.client_id, 'end', iid=position['tradingsymbol'] + '#' + account.client_id + '#' + position['product'] + '#' +position['exchange'],
                                                        tags=('content',),
                                                        values=(
                                position['tradingsymbol'],
                                position['product'],
                                position['quantity'],
                                'Add',
                                'Exit'))

            # pol += 1

        treeScroll = ttk.Scrollbar(self.openPositionsFrame)
        treeScroll.configure(command=(self.openPositionsTree.yview))
        self.openPositionsTree.configure(yscrollcommand=(treeScroll.set))
        self.openPositionsTree.tag_configure('order', background='#ecf2fe', font=(
            None, 11))
        self.openPositionsTree.tag_configure('content', font=(
            None, 9))
        self.openPositionsTree.bind(
            '<ButtonRelease-1>', lambda event: self.open_positions_tree_click_event(event, 'positions'))
        self.openPositionsTree.pack(side=LEFT, fill=BOTH, anchor='ne')
        treeScroll.pack(side=RIGHT, fill=Y)

    def refresh_orders(self):
        list_of_orders = []
        is_account_logged_in = False
        for account in self.listOfAccounts:
            if account.authStatus == 'Logged In':
                is_account_logged_in = True
                try:
                    orders = account.get_orders()
                    if orders != None and type(orders) is dict and orders['status'] == 'success':
                        if len(orders['data']) > 0:
                            for order in orders['data']:
                                list_of_orders.append(order)
                except Exception as e:
                    print(e,'refresh_orders >>>')
        if is_account_logged_in == False:
            showerror('Refresh Orders','No account logged in for refresh',parent=self.view_order_win)
            return
        print('\n>>>>>>>>>',list_of_orders,'\n>>>>>>>>>')
        # if len(list_of_orders) != 0 :
        #     for order in list_of_orders:
        #         print(order['order_id'],order['placed_by'],order['status'],order['quantity'],
        #         order['filled_quantity'],'\n')
        try : 
            list_of_orders_tuples = self.order_db.get_orders()
            if len(list_of_orders_tuples) > 0 and len(list_of_orders) != 0:
                for order in list_of_orders:
                    for internal_order in list_of_orders_tuples:
                        try:
                            internal_order_object = json.loads(internal_order[2])
                        
                            # print(order != None and type(order) is dict and (order['placed_by'] in internal_order_object) and internal_order_object[order['placed_by']]['data']['order_id'] != None and internal_order_object[order['placed_by']]['data']['order_id'] == order['order_id'],'820')
                            
                            if order != None and type(order) is dict and (order['placed_by'] in internal_order_object) and internal_order_object[order['placed_by']]['data']['order_id'] != None and internal_order_object[order['placed_by']]['data']['order_id'] == order['order_id']:
                                print(json.loads(internal_order[2]),'823')
                                self.update_order_status(
                                                            internal_order[0], order['placed_by'], order['status'],'exchange_order_status',False)
                                self.update_order_status(
                                                            internal_order[0], order['placed_by'], order['filled_quantity'],'fill_quantity',False)
                            if order != None and type(order) is dict and (order['placed_by'] in internal_order_object) and internal_order_object[order['placed_by']]['data']['order_id'] != None and internal_order_object[order['placed_by']]['data']['order_id'] == order['parent_order_id']:
                                print(json.loads(internal_order[2]),'823')
                                self.update_order_status(
                                                            internal_order[0], order['placed_by'], 'COMPLETE' if order['meta']['iceberg']['remaining_quantity'] == 0 else 'PARTIALLY FILLED'
                                                            ,'exchange_order_status',False)
                                self.update_order_status(
                                                            internal_order[0], order['placed_by'], order['meta']['iceberg']['total_quantity'] - order['meta']['iceberg']['remaining_quantity'],'fill_quantity',False)
                        except Exception as e:
                            print(e,'824',order,internal_order_object)
        except Exception as e:
            print(e,'831')
        self.recreate_running_orders_tree()

    def recreate_running_orders_tree(self):
        if hasattr(self,'runningOrdersTree') == False:
            return
        list_of_orders_tuples = self.order_db.get_orders()
        if len(self.runningOrdersTree.get_children()) != 0:
            self.runningOrdersTree.delete(
                *self.runningOrdersTree.get_children())

        pos = 1
        for item in list_of_orders_tuples:
            loaded_order = json.loads(item[1])
            account_level_orders = json.loads(item[2])
            self.runningOrdersTree.insert('', 'end', text=pos, iid=item[0],
                                          values=(
                                              item[0],
                                              loaded_order['tradingsymbol'], loaded_order['transactiontype'],
                                              loaded_order['quantity']), open=True, tags=('order',))
            pos += 1
            for key in account_level_orders.keys():
                self.runningOrdersTree.insert(item[0], 'end', iid=item[0] + '#' + key,
                                              values=(
                    key,
                    account_level_orders[key]['exchange_order_status'] if account_level_orders[key].get(
                        'exchange_order_status') else 'UNKOWN',
                    '{0}/{1}'.format(account_level_orders[key]['fill_quantity'],account_level_orders[key]['order_quantity']) if account_level_orders[key].get(
                        'fill_quantity') else 0,
                    # 'Unknown'
                    'Modify',
                    'Delete'), open=True, tags=('content',account_level_orders[key]['exchange_order_status'] if account_level_orders[key].get(
                        'exchange_order_status') else 'UNKOWN'))

    def open_positions_tree_click_event(self, event, tree_type):
        curItem = ''
        col = ''
        if tree_type == 'positions':
            curItem = self.openPositionsTree.item(
                self.openPositionsTree.focus())
            col = self.openPositionsTree.identify_column(event.x)
            row = self.openPositionsTree.identify_row(event.y)
        else:
            curItem = self.runningOrdersTree.item(
                self.runningOrdersTree.focus())
            col = self.runningOrdersTree.identify_column(event.x)
            row = self.runningOrdersTree.identify_row(event.y)
        print('curItem = ', curItem)
        print('col = ', col.split('#'))
        print('row = ', row.split('#'))
        column_pure_number = int(col.split('#')[1])

        if len(curItem['values']) >= column_pure_number:
            print(curItem['values'][column_pure_number - 1])
            selected_cell = curItem['values'][column_pure_number - 1]
            print(selected_cell, '794 selected')
            row_content = row.split('#')

            if selected_cell == 'Delete':
                if curItem['values'][1] == 'CANCELLED' or curItem['values'][1] == 'COMPLETE' or curItem['values'][1] == 'REJECTED':
                    if showerror('Copy Trader', "Order can't be deleted since its already processed", parent=self.view_order_win) == 'ok':
                        return

                try:
                    exg_order_id = self.find_exchange_order_id(
                        row_content[0], row_content[1])
                    self.delete_order_single(
                        exg_order_id, row_content[1], row_content[0])
                except Exception as e:
                    print(e)
                pass
            elif selected_cell == 'Modify':
                if curItem['values'][1] == 'CANCELLED' or curItem['values'][1] == 'COMPLETE' or curItem['values'][1] == 'REJECTED':
                    if showerror('Copy Trader', "Order can't be modified since its already processed", parent=self.view_order_win) == 'ok':
                        return
                try:
                    exg_order_id = self.find_exchange_order_id(
                        row_content[0], row_content[1])
                    self.modify_order_single(exg_order_id, row_content[1],row_content[0])
                except Exception as e:
                    print(e)
            elif selected_cell == 'Exit' and abs(int(curItem['values'][2])) != 0:
                self.exit_position_single(row_content[1], curItem['values'])
                pass
            elif selected_cell == 'Add' and abs(int(curItem['values'][2])) != 0:
                self.exit_position_single(row_content[1], curItem['values'],'add')
                pass

    def multiply_instrument_lots(self):
        if self.exit_quant.get() != '' and self.exit_lot.get() != '':
            print(type(self.exit_quant.get()), type(self.exit_lot.get()))
            self.exit_total_quant.configure(state='normal')
            self.exit_total_quant.delete(0, END)
            self.exit_total_quant.insert(0, int(float(self.exit_quant.get()))
                                         * int(float(self.exit_lot.get())))
            self.exit_total_quant.configure(state='disable')

    def multiply_instrument_lots_generic(self, exit_quant, exit_total_quant, exit_lot, keychar):
        if exit_quant.get() != '' and exit_lot.get() != '':
            print(exit_quant.get(), exit_lot.get(), exit_total_quant)
            exit_total_quant.configure(state='normal')
            exit_total_quant.delete(0, END)
            exit_total_quant.insert(0, int(float(exit_quant.get()))
                                    * int(float(exit_lot.get())))
            exit_total_quant.configure(state='disable')
            print(int(float(exit_quant.get()))
                  * int(float(exit_lot.get())), exit_total_quant.get())
    
    def calculate_iceberg_lots_generic(self,variety):
        quant = int(float(self.exit_lot.get()))
        if variety.get() == 'iceberg':
            if quant < 5:
                quant = 5
            self.exit_iceberg_quantity.delete(0,END)
            self.exit_iceberg_quantity.insert(0,ceil(quant/int(self.exit_iceberg_lot.get())) * int(self.exit_quant.get()))
        else:
            self.exit_iceberg_lot.delete(0,END)
            self.exit_iceberg_quantity.delete(0,END)
            self.exit_iceberg_lot.insert(0,0)
            self.exit_iceberg_quantity.insert(0,0)
    def exit_position_single(self, parent_tree_id, trade_values_array,action_type = 'exit'):
        
        print(self.openPositionsTree.get_children(parent_tree_id), 'MMMMMMMMMM')
        if self.is_singleorder_exit_win_initial_load == False:
            self.single_order_exit_win.destroy()

        if self.is_singleorder_exit_win_initial_load == True:
            self.is_singleorder_exit_win_initial_load = False

        self.single_order_exit_win = Toplevel(self, padx=10, width=250, height=190,
                                              pady=10)
        window_text = 'Exit'
        if action_type != 'exit':
            window_text = 'Add'
        self.single_order_exit_win.title(window_text + ' ' + trade_values_array[0])

        try:
            instrument_response = []
            instrument_response = self.zerodha_db.get_specific_instruments_data_by_tradingsymbol(
                trade_values_array[0])
            print(instrument_response, 'instrument_response')
            if len(instrument_response) == 0:
                print(instrument_response)
                if showerror('Copy Trader', 'Instrument does not exist') == 'ok':
                    self.single_order_exit_win.destroy()
                    return
        except Exception as e:
            print(e, '495 ltp_zerodha')
            return 0

        self.exit_ordertype = StringVar()
        self.exit_ordertype.set('LIMIT')
        self.exit_varietytype = StringVar()
        self.exit_varietytype.set('regular')
        self.exit_producttype = StringVar()

        self.exit_transactiontype = StringVar()
        if action_type == 'exit':
            if int(trade_values_array[2]) < 0:
                self.exit_transactiontype.set('BUY')
            else:
                self.exit_transactiontype.set('SELL')
        else:
            if int(trade_values_array[2]) < 0:
                self.exit_transactiontype.set('SELL')
            else:
                self.exit_transactiontype.set('BUY')

            
        buySellFrame = LabelFrame(
            self.single_order_exit_win, height=300, width=100, text='Buy/Sell')
        buySellFrame.pack(side=TOP, fill=X)
        radio1 = Radiobutton(buySellFrame, text='Buy', command=(
            lambda: self.orderType()), variable=(self.exit_transactiontype), value='BUY')
        radio1.pack(side=LEFT)
        radio1.config(state='disable')
        radio2 = Radiobutton(buySellFrame, text='Sell', command=(
            lambda: self.orderType()), variable=(self.exit_transactiontype), value='SELL')
        radio2.pack(side=LEFT)
        radio2.config(state='disable')

        # CNC,NRML,MIS
        if trade_values_array[1] == 'NRML':
            self.exit_producttype.set('NRML')
        elif trade_values_array[1] == 'MIS':
            self.exit_producttype.set('MIS')
        else:
            self.exit_producttype.set('CNC')

        order_type = LabelFrame(
            self.single_order_exit_win, text='Product type')
        rad1 = Radiobutton(order_type, text='Market', command=(
            lambda: self.orderType()), variable=(self.exit_ordertype), value='MARKET')
        rad1.pack(side=LEFT, fill=NONE)
        rad2 = Radiobutton(order_type, text='Limit', command=(
            lambda: self.orderType()), variable=(self.exit_ordertype), value='LIMIT')
        rad2.pack(side=LEFT, fill=NONE)
        order_type.pack(side=TOP, fill=X)

        variety_type = LabelFrame(
            self.single_order_exit_win, text='Variety')
        rad1 = Radiobutton(variety_type, text='Regular', command=self.is_exit_iceberg_logic, 
        variable=(self.exit_varietytype), value='regular')
        rad1.pack(side=LEFT, fill=NONE)
        rad2 = Radiobutton(variety_type, text='Iceberg', command=self.is_exit_iceberg_logic, 
        variable=(self.exit_varietytype), value='iceberg')
        rad2.pack(side=LEFT, fill=NONE)
        variety_type.pack(side=TOP, fill=X)



        producttype = LabelFrame(self.single_order_exit_win,
                                 height=300, width=400, text='Product type')
        producttype.pack(side=TOP, fill=X)
        rad1 = Radiobutton(producttype, text='CNC', command=(
            lambda: self.orderType()), variable=(self.exit_producttype), value='CNC')
        rad1.pack(side=LEFT)
        rad2 = Radiobutton(producttype, text='NRML', command=(
            lambda: self.orderType()), variable=(self.exit_producttype), value='NRML')
        rad2.pack(side=LEFT)
        rad3 = Radiobutton(producttype, text='MIS', command=(
            lambda: self.orderType()), variable=(self.exit_producttype), value='MIS')
        rad3.pack(side=LEFT)

        price_combo1 = Frame(self.single_order_exit_win,
                             height=300, width=100, padx=1, pady=5)
        labP = Label(price_combo1, width=10, text='Price')

        self.exit_price = Entry(price_combo1, width=10, text='Enter Price')
        self.exit_price.delete(0, END)
        self.exit_price.insert(
            0, float(self.get_last_traded_price(trade_values_array[0], False)))

        labP.pack(side=LEFT)
        self.exit_price.pack(side=LEFT)


        labIL = Label(price_combo1, width=5, text='Legs')

        self.exit_iceberg_lot = Entry(price_combo1, width=5, text='Legs')
        labIL.pack(side=LEFT)
        self.exit_iceberg_lot.pack(side=LEFT)
        self.exit_iceberg_lot.bind(
            '<KeyRelease>', (lambda e: self.calculate_iceberg_lots_generic(self.exit_varietytype)))

        labIQ = Label(price_combo1, width=15, text='Iceberg Quantity')
        self.exit_iceberg_quantity = Entry(price_combo1, width=10, text='Iceberg Quantity')
        labIQ.pack(side=LEFT)
        self.exit_iceberg_quantity.pack(side=LEFT)


        price_combo1.pack(side=TOP, fill=X)
        price_combo2 = Frame(self.single_order_exit_win,
                             height=300, width=100, padx=1, pady=5)
        labQ = Label(price_combo2, width=10, text='Quantity')
        self.exit_quant = Entry(price_combo2, width=10, text='Enter Quantity')
        self.exit_quant.delete(0, END)
        self.exit_quant.insert(0, int(instrument_response[0][8]))
        labQ.pack(side=LEFT)
        self.exit_quant.pack(side=LEFT)
        self.exit_quant.configure(state='disable')


        labM = Label(price_combo2, width=7, text='Lot')

        self.exit_lot = Spinbox(price_combo2, from_=0, to=50, values=(0, 10, 20, 30, 40, 50),
                                width=5, wrap=False,
                                command=(lambda: self.multiply_instrument_lots()))
        self.exit_lot.delete(0, END)
        self.exit_lot.insert(
            0, abs(int(trade_values_array[2]))/int(instrument_response[0][8]))
        def combineFunctions(self):
            self.calculate_iceberg_lots_generic(self.exit_varietytype)
            self.multiply_instrument_lots()
        self.exit_lot.bind(
            '<KeyRelease>', (lambda e: combineFunctions(self)))
        labM.pack(side=LEFT)
        self.exit_lot.pack(side=LEFT)

        labTQ = Label(price_combo2, width=15, text='Total Quantity')
        self.exit_total_quant = Entry(
            price_combo2, width=10, text='Enter Total Quantity')
        self.exit_total_quant.delete(0, END)
        self.exit_total_quant.insert(
            0, int(float(self.exit_lot.get())) * int(float(self.exit_quant.get())))
        # self.exit_total_quant.configure(state='disable')
        labTQ.pack(side=LEFT)
        self.exit_total_quant.pack(side=LEFT)
        price_combo2.pack(side=TOP, fill=X)
        btn_text = 'Exit Position'
        if action_type != 'exit':
            btn_text = 'Add Position'
        exit_object = {
            'tradingsymbol': trade_values_array[0],
                       'exchange': instrument_response[0][11],
                       'transaction_type': self.exit_transactiontype.get(),
                       'order_type': self.exit_ordertype.get(),
                       'quantity': int(float(self.exit_total_quant.get())),
                       'product': self.exit_producttype.get(),
                       'price': float(self.exit_price.get()),
                       'validity': 'DAY',
                       'variety': 'regular'
        }
        if self.exit_iceberg_quantity.get() != '' and self.exit_iceberg_quantity.get() != None and int(self.exit_iceberg_quantity.get()) >= 5 and self.exit_iceberg_lot.get() != '' and self.exit_iceberg_lot.get() != None:
            exit_object['variety'] = 'iceberg'
            exit_object['iceberg_quantity'] = int(self.exit_iceberg_quantity.get())
            exit_object['iceberg_legs'] = int(self.exit_iceberg_lot.get())
        Button(self.single_order_exit_win, text=btn_text, width=15,
               command=(lambda: self.execute_exit_position(
                   parent_tree_id,
                   exit_object
               ))).pack(side=TOP)

    def is_exit_iceberg_logic(self):
        self.exit_iceberg_lot.delete(0,END)
        self.exit_iceberg_quantity.delete(0,END)
        if self.exit_varietytype.get() == 'iceberg':
            print(self.exit_lot.get(),self.exit_quant.get())
            if int(float(self.exit_lot.get())) >= 5:
                self.exit_iceberg_lot.insert(0,3)
                self.exit_iceberg_quantity.insert(0, ceil(int(float(self.exit_lot.get())/3)) * int(self.exit_quant.get()))
        else:
            self.exit_iceberg_lot.insert(0,0)
            self.exit_iceberg_quantity.insert(0, 0)
    def modify_multiple(self):
        if hasattr(self,'execute_multiple_modify_win'):
            self.execute_multiple_modify_win.destroy()

        if len(self.runningOrdersTree.selection()) == 0:
            if showwarning('Copy Trader', 'No orders selected for deletion', parent=self.view_order_win) == 'ok':
                return
        selected_orders = {}
        accounts_by_order_id = {}
        
        for item in self.runningOrdersTree.selection():
            print(item.split('#'),type(item.split('#')) is list and len(item.split('#')) > 1)   
            # if 'COMPLETE' , CANCELLED, and REJECTED.  
            if type(item.split('#')) is list and len(item.split('#')) > 1:
                if item.split('#')[0] not in accounts_by_order_id:
                    accounts_by_order_id[item.split('#')[0]] = []
                
                accounts_by_order_id[item.split('#')[0]].append(item.split('#')[1])
        print(accounts_by_order_id)       
        self.execute_multiple_modify_win = Toplevel(self, padx=10, width=815, height=590,
                                                       pady=10)
            # win.config()
        self.execute_multiple_modify_win.pack_propagate(0)
        delete_button_frame = Frame(
            self.execute_multiple_modify_win, pady=15)
        delete_button_frame.pack(side=TOP, fill=X)
        execute_delete_button = Button(delete_button_frame, text='Modify Orders')

        execute_delete_button.pack(side=LEFT, fill=NONE, anchor='e')
        canvas = Canvas(self.execute_multiple_modify_win, borderwidth=0,
                        background="#e0e0e0", width=815)
        executionFrame = Frame(
            canvas, height=300, padx=2)
        executionFrame.pack(side=TOP, fill=BOTH)

        sbar = Scrollbar(self.execute_multiple_modify_win,
                            orient="vertical", command=(canvas.yview))
        canvas.configure(yscrollcommand=sbar.set)
        sbar.pack(side=RIGHT, fill=Y)
        canvas.pack(side=LEFT, fill=BOTH)
        canvas.create_window((1, 1), window=executionFrame, anchor="nw")
        executionFrame.bind(
            "<Configure>", lambda event, canvas=canvas: self.canvas_frame_reconfigurer(canvas))
        exchange_order_ids_by_account = {}
        self.multiple_mod_form_object_consolidated = {}
        self.multiple_mod_form_object_price_consolidated = {}
        self.multiple_mod_form_object_trigger_price_consolidated = {}
        self.multiple_mod_form_object_order_variety_consolidated = {}
        if len(accounts_by_order_id.keys()) > 0:
            for item in accounts_by_order_id.keys():
                if item not in exchange_order_ids_by_account:
                    exchange_order_ids_by_account[item] = {}
                list_of_orders = self.order_db.get_order_by_timestamp(item)              
                if len(list_of_orders) > 0:
                    loaded_order_json = json.loads(
                        list_of_orders[0][2])
                    trading_symbol = json.loads(
                        list_of_orders[0][1])['tradingsymbol']
                    label_frame = LabelFrame(executionFrame,text=item + '\t\t' + trading_symbol)
                    self.multiple_mod_form_object_price_consolidated[item] = DoubleVar()
                    self.multiple_mod_form_object_trigger_price_consolidated[item] = DoubleVar()
                    price_frame = Frame(label_frame,padx=10)
                    price_frame.pack(side=TOP,fill=X)
                    Label(price_frame,text='Price').pack(side=LEFT,fill=NONE)
                    Entry(price_frame,textvariable=self.multiple_mod_form_object_price_consolidated[item],width=10).pack(side=LEFT,fill=NONE)
                    Label(price_frame,text='Trigger Price').pack(side=LEFT,fill=NONE)
                    Entry(price_frame,textvariable=self.multiple_mod_form_object_trigger_price_consolidated[item],width=10).pack(side=LEFT,fill=NONE)

                    check_var = StringVar()
                    """ Label(price_frame,text='Limit').pack(side=LEFT,fill=NONE)
                    Checkbutton(price_frame,variable=check_var,command=lambda key = item,state = check_var:self.toggleLimitMarket(key,state)).pack(side=LEFT,fill=NONE)
                    Label(price_frame,text='SL').pack(side=LEFT,fill=NONE)
                    Checkbutton(price_frame,variable=check_var,command=lambda key = item,state = check_var:self.toggleLimitMarket(key,state)).pack(side=LEFT,fill=NONE) """
                    Radiobutton(price_frame,text='Limit',variable=check_var,value='LIMIT',command=lambda key = item,state = check_var:self.toggleLimitMarket(key,state)).pack(side=LEFT,fill=NONE)
                    Radiobutton(price_frame,text='Market',variable=check_var,value='MARKET',command=lambda key = item,state = check_var:self.toggleLimitMarket(key,state)).pack(side=LEFT,fill=NONE)
                    Radiobutton(price_frame,text='SL',variable=check_var,value='SL',command=lambda key = item,state = check_var:self.toggleLimitMarket(key,state)).pack(side=LEFT,fill=NONE)

                    for order in accounts_by_order_id[item]:
                        if order in loaded_order_json and loaded_order_json[order]['exchange_order_status'] != 'COMPLETE' and loaded_order_json[order]['exchange_order_status'] != 'REJECTED' and loaded_order_json[order]['exchange_order_status'] != 'CANCELLED':
                            exchange_order_ids_by_account[item][order] = loaded_order_json[order]['data']['order_id']
                            order_frame = Frame(label_frame,padx=10,pady=5)
                            Label(order_frame,text=order + '\t\t',anchor='w').pack(side=LEFT,fill=NONE)
                            account_hash = item+'#'+order+'#'+loaded_order_json[order]['data']['order_id']
                            try:

                                self.multiple_mod_form_object_consolidated[account_hash] = {
                                        'quantity': IntVar(value=int(loaded_order_json[order]['order_quantity'])),
                                        'price':IntVar(value=0),
                                        'order_type':StringVar(value='LIMIT'),
                                        'trading_symbol':trading_symbol,
                                        'trigger_price':DoubleVar()
                                }
                            except Exception as e:
                                self.multiple_mod_form_object_consolidated[account_hash] = {
                                        'quantity': IntVar(value=0),
                                        'price':IntVar(value=0),
                                        'order_type':StringVar(value='LIMIT'),
                                        'trading_symbol':trading_symbol,
                                        'trigger_price':DoubleVar()
                                }
                                print(e)
                            Radiobutton(order_frame,text='SL',variable=self.multiple_mod_form_object_consolidated[account_hash]['order_type'],value='SL').pack(side=RIGHT,fill=NONE)
                            Radiobutton(order_frame,text='Market',variable=self.multiple_mod_form_object_consolidated[account_hash]['order_type'],value='MARKET').pack(side=RIGHT,fill=NONE)
                            Radiobutton(order_frame,text='Limit',variable=self.multiple_mod_form_object_consolidated[account_hash]['order_type'],value='LIMIT').pack(side=RIGHT,fill=NONE)

                            if 'trigger_price' in loaded_order_json[order]:
                               self.multiple_mod_form_object_consolidated[account_hash]['order_type'].set('SL')
                               self.multiple_mod_form_object_consolidated[account_hash]['trigger_price'].set(loaded_order_json[order]['trigger_price'])
                            
                            quantEntry = Entry(order_frame,textvariable=self.multiple_mod_form_object_consolidated[account_hash]['quantity'])
                            quantEntry.pack(side=RIGHT,fill=NONE)
                            if 'variety' in loaded_order_json[order] and loaded_order_json[order]['variety'] == 'iceberg':
                                quantEntry.configure(state='disabled')
                                self.multiple_mod_form_object_order_variety_consolidated[account_hash] = 'iceberg'
                            else:
                                self.multiple_mod_form_object_order_variety_consolidated[account_hash] = 'regular'
                            order_frame.pack(side=TOP,fill=X)
                    label_frame.pack(side=TOP,fill=X)
                    print(loaded_order_json,'\n')
        print(exchange_order_ids_by_account)
        if len(exchange_order_ids_by_account.values()) == 0:
            if showwarning('Copy Trader', 'All orders have been processed', parent=self.execute_multiple_modify_win) == 'ok':
                self.execute_multiple_modify_win.destroy()
                return
        execute_delete_button.configure(command=lambda delete_json=exchange_order_ids_by_account:
        self.run_multiple_modifications(delete_json))

    def run_multiple_modifications(self,executejson):
        post_order_screens = {}
        for key in self.multiple_mod_form_object_consolidated.keys():
            print(key.split('#'))
            exchange_order_id = key.split('#')[2]
            account_number = key.split('#')[1]
            order_id = key.split('#')[0]
            order_object = {
                'price':self.multiple_mod_form_object_price_consolidated[key.split('#')[0]].get(),
                'quantity':self.multiple_mod_form_object_consolidated[key]['quantity'].get(),
                'order_type':self.multiple_mod_form_object_consolidated[key]['order_type'].get(),
                'trigger_price': self.multiple_mod_form_object_trigger_price_consolidated[key.split('#')[0]].get()
            }
            if self.multiple_mod_form_object_order_variety_consolidated[key] == 'iceberg':
                del order_object['quantity']
            print(order_object,account_number,exchange_order_id,'\n')
            for acc in self.listOfAccounts:
                    # print(acc.client_id,account_number,'delete_order_single')
                    if acc.client_id == account_number and acc.authStatus == 'Logged In':
                        print('modify_order', acc)
                        is_account_logged_in = True
                        try:
                            mod_response = acc.modify_order(
                                exchange_order_id, order_object, 
                                self.multiple_mod_form_object_order_variety_consolidated[key])
                            post_order_screens[account_number + ' ' + self.multiple_mod_form_object_consolidated[key]['trading_symbol']] = mod_response
                            if mod_response != None and mod_response['status'] == 'success':
                                if 'quantity' in order_object:
                                    self.update_order_status(
                                                        order_id, account_number,float(order_object['quantity']) ,'order_quantity')
                                if 'trigger_price' in order_object:
                                    self.update_order_status(
                                                        order_id, account_number,float(order_object['trigger_price']) ,'trigger_price')
                        except Exception as e:
                            print(e)
                            post_order_screens[account_number + ' ' + self.multiple_mod_form_object_consolidated[key]['trading_symbol']] = e
        self.post_order_execeution_screen(post_order_screens,'Multiple Modifications')                            
        if hasattr(self,'execute_multiple_modify_win'):
            self.execute_multiple_modify_win.destroy()

    def toggleLimitMarket(self,order_id,state):
        print(state.get())
        for item in self.multiple_mod_form_object_consolidated.keys():
            if item.split('#')[0] == order_id:
                print(state.get())
                if state.get() == 'SL':
                    self.multiple_mod_form_object_consolidated[item]['order_type'].set('SL')
                elif state.get() == 'MARKET':
                    self.multiple_mod_form_object_consolidated[item]['order_type'].set('MARKET')
                elif state.get() == 'LIMIT':
                    self.multiple_mod_form_object_consolidated[item]['order_type'].set('LIMIT')
    def delete_order_history(self):
        if self.order_db.delete_all_orders():
            self.recreate_running_orders_tree()
            if showinfo('Copy Trader','All orders deleted') == 'ok':
                pass

    def delete_multiple(self):
        if hasattr(self,'execute_multiple_delete_win'):
            self.execute_multiple_delete_win.destroy()
        if len(self.runningOrdersTree.selection()) == 0:
            if showwarning('Copy Trader', 'No orders selected for deletion', parent=self.view_order_win) == 'ok':
                return
        selected_orders = {}
        accounts_by_order_id = {}
        for item in self.runningOrdersTree.selection():
            print(item.split('#'),type(item.split('#')) is list and len(item.split('#')) > 1)   
            # if 'COMPLETE' , CANCELLED, and REJECTED.  
            if type(item.split('#')) is list and len(item.split('#')) > 1:
                if item.split('#')[0] not in accounts_by_order_id:
                    accounts_by_order_id[item.split('#')[0]] = []
                
                accounts_by_order_id[item.split('#')[0]].append(item.split('#')[1])
        print(accounts_by_order_id)       
        self.execute_multiple_delete_win = Toplevel(self, padx=10, width=815, height=590,
                                                       pady=10)
            # win.config()
        self.execute_multiple_delete_win.pack_propagate(0)
        delete_button_frame = Frame(
            self.execute_multiple_delete_win, pady=15)
        delete_button_frame.pack(side=TOP, fill=X)
        execute_delete_button = Button(delete_button_frame, text='Delete Orders')

        execute_delete_button.pack(side=LEFT, fill=NONE, anchor='e')
        canvas = Canvas(self.execute_multiple_delete_win, borderwidth=0,
                        background="#e0e0e0", width=815)
        executionFrame = Frame(
            canvas, height=300, padx=2)
        executionFrame.pack(side=TOP, fill=BOTH)

        sbar = Scrollbar(self.execute_multiple_delete_win,
                            orient="vertical", command=(canvas.yview))
        canvas.configure(yscrollcommand=sbar.set)
        sbar.pack(side=RIGHT, fill=Y)
        canvas.pack(side=LEFT, fill=BOTH)
        canvas.create_window((1, 1), window=executionFrame, anchor="nw")
        executionFrame.bind(
            "<Configure>", lambda event, canvas=canvas: self.canvas_frame_reconfigurer(canvas))
        exchange_order_ids_by_account = {}
        if len(accounts_by_order_id.keys()) > 0:
            for item in accounts_by_order_id.keys():
                if item not in exchange_order_ids_by_account:
                    exchange_order_ids_by_account[item] = {}
                list_of_orders = self.order_db.get_order_by_timestamp(item)              
                if len(list_of_orders) > 0:
                    loaded_order_json = json.loads(
                        list_of_orders[0][2])
                    trading_symbol = json.loads(
                        list_of_orders[0][1])
                    trading_symbol = trading_symbol['tradingsymbol']
                    label_frame = LabelFrame(executionFrame,text=item + '\t\t' + trading_symbol)
                    for order in accounts_by_order_id[item]:
                        if order in loaded_order_json and loaded_order_json[order]['exchange_order_status'] != 'COMPLETE' and loaded_order_json[order]['exchange_order_status'] != 'REJECTED' and loaded_order_json[order]['exchange_order_status'] != 'CANCELLED':
                            exchange_order_ids_by_account[item][order] = loaded_order_json[order]['data']['order_id']
                            Label(label_frame,text=order + '\t\t'+ str(loaded_order_json[order]['order_quantity']),anchor='w',padx=10,pady=5).pack(side=TOP,fill=X)
                    label_frame.pack(side=TOP,fill=X)
                    print(loaded_order_json,'\n')
        print(exchange_order_ids_by_account)
        if len(exchange_order_ids_by_account.values()) == 0:
            if showwarning('Copy Trader', 'All orders have been processed', parent=self.execute_multiple_delete_win) == 'ok':
                self.execute_multiple_delete_win.destroy()
                return
            
        execute_delete_button.configure(command= lambda delete_json=exchange_order_ids_by_account:
        self.run_multiple_delete(delete_json))
    def run_multiple_delete(self,delete_json):
        post_order_screen = {}
        for order_id in delete_json.keys():
            for account_id in delete_json[order_id]:
                for acc in self.listOfAccounts:
                    if acc.client_id == account_id and acc.authStatus == 'Logged In':
                        # print('delete_order_single', acc)
                        print('delete_order_single',account_id,delete_json[order_id][account_id])

                        is_account_logged_in = True
                        try:
                            deletion_response = acc.cancel_order(delete_json[order_id][account_id])
                            # showinfo('Cancel Order', deletion_response,
                            #         parent=self.view_order_win)
                            if deletion_response != None and deletion_response['status'] == 'success':
                                self.update_order_status(
                                    order_id, account_id, 'CANCELLED','exchange_order_status')
                            post_order_screen[account_id + ' ' + delete_json[order_id][account_id]] = deletion_response
                        except Exception as e:
                            print(e)
                            post_order_screen[account_id + ' ' + delete_json[order_id][account_id]] = e
                            # showerror(
                            #     'Copy Trader', 'Something went wrong please try again later', parent=self.view_order_win)
                    # else:
                    #     post_order_screen[account_id + ' ' + delete_json[order_id][account_id]] = 'Not logged in'
                    #     pass
        self.post_order_execeution_screen(post_order_screen, 'Order Screen')

    def excecute_instruments_multiple(self, action_type='exit'):
        # print(self.openPositionsTree.selection())
        if len(self.openPositionsTree.selection()) == 0:

            if showwarning('Copy Trader', 'No positions selected for execution', parent=self.view_order_win) == 'ok':
                return
        instrument_item = self.openPositionsTree.selection()[0]
        if type(instrument_item.split('#')) is list and len(instrument_item.split('#')) == 1:
            return
        processed_positions = []
        processed_positions_by_instrument = {}
        # complete_tree = []
        selection_tree = []
        selected_instrument_for_check = instrument_item.split('#')[0]
        for item in self.openPositionsTree.get_children():
            print(instrument_item.split('#'), item, 'ERF',
                  self.openPositionsTree.get_children(item), 'openPositionsTree \n')
            for child_items in self.openPositionsTree.get_children(item):
                if selected_instrument_for_check == child_items.split('#')[0]:
                    selection_tree.append(child_items)
        # for item in self.openPositionsTree.selection():
        for item in selection_tree:

            # print(type(item.split('#')) is list and len(item.split('#')) > 1)
            processed_arr = item.split('#')
            if type(processed_arr) is list and len(processed_arr) > 1:
                print(processed_arr, self.openPositionsTree.item(item))
                processed_positions.append(
                    {'account': processed_arr[1], 'position': processed_arr[0], 'values': self.openPositionsTree.item(item)['values']})
                if processed_positions_by_instrument.get(processed_arr[0]):
                    processed_positions_by_instrument[processed_arr[0]].append({
                        'account': processed_arr[1], 'values': self.openPositionsTree.item(item)['values']
                    })
                else:
                    processed_positions_by_instrument[processed_arr[0]] = []
                    processed_positions_by_instrument[processed_arr[0]].append({
                        'account': processed_arr[1], 'values': self.openPositionsTree.item(item)['values']
                    })
        if len(processed_positions) == 0:
            if showwarning('Copy Trader', 'No positions selected for execution', parent=self.view_order_win) == 'ok':
                return
        print(processed_positions_by_instrument)

        form_object_temp = {}
        form_object_price = {}
        for key, values in processed_positions_by_instrument.items():
            # print(key,values,'processed_positions_by_instrument.items()\n')
            instrument_response = []

            try:
                instrument_response = self.zerodha_db.get_specific_instruments_data_by_tradingsymbol(
                    key)
                print(instrument_response)

            except Exception as e:
                print(e, '495 ltp_zerodha')
                instrument_response = []
                return 0

            ltp = float(self.get_last_traded_price(key, False))
            if form_object_price.get(key) == None:
                        # form_object_price[key] = DoubleVar()
                        form_object_price[key] = ltp
            for item in values:
                if int(item['values'][2]) != 0:
                    if len(instrument_response) == 0:
                        # print(instrument_response)
                        if showerror('Copy Trader', 'Instrument does not exist') == 'ok':
                            self.single_order_exit_win.destroy()
                            return
                    # print(self.form_object_exit.get(key))
                    if form_object_temp.get(key) == None:
                        form_object_temp[key] = []

                    form_object_temp[key].append({
                        'exit_ordertype': 'LIMIT',
                        'exit_producttype': '',
                        'exit_transactiontype': '',
                        'quantity': instrument_response[0][8],
                        'lots': int(abs(int(item['values'][2]))/int(instrument_response[0][8])),
                        'ltp': ltp,
                        'instrument': item['values'][0],
                        'account': item['account'],
                        'total_quantity': item['values'][2],
                        'exchange': instrument_response[0][-1],
                        'iceberg_legs': 0,
                        'iceberg_quantity': 0
                    })
                    if action_type == 'exit':
                        if int(item['values'][2]) < 0:
                            form_object_temp[key][- 1]['exit_transactiontype'] = 'BUY'
                        else:
                            form_object_temp[key][- 1]['exit_transactiontype'] = 'SELL'
                    else:
                        if int(item['values'][2]) < 0:
                            form_object_temp[key][- 1]['exit_transactiontype'] = 'SELL'
                        else:
                            form_object_temp[key][-1]['exit_transactiontype'] = 'BUY'
                    # print(temp_var['exit_transactiontype'].get(),'1055')
                    if item['values'][1] == 'NRML':
                        form_object_temp[key][-1]['exit_producttype'] = 'NRML'
                    elif item['values'][1] == 'MIS':
                        form_object_temp[key][-1]['exit_producttype'] = 'MIS'
                    else:
                        form_object_temp[key][-1]['exit_producttype'] = 'CNC'

                    # print(temp_var['exit_producttype'].get(),'1055')
                    # form_object[key].append(temp_var)
            # print(len(self.form_object_exit))
        if len(form_object_temp) > 0:

            self.execute_multiple_order_win = Toplevel(self, padx=10, width=1015, height=590,
                                                       pady=10)
            # win.config()
            self.execute_multiple_order_win.pack_propagate(0)
            execute_button_frame = Frame(
                self.execute_multiple_order_win, pady=15)
            execute_button_frame.pack(side=TOP, fill=X)
            execute_exit_button = Button(execute_button_frame, text=str(action_type).capitalize())

            execute_exit_button.pack(side=LEFT, fill=NONE, anchor='e')
            canvas = Canvas(self.execute_multiple_order_win, borderwidth=0,
                            background="#e0e0e0", width=975)
            executionFrame = Frame(
                canvas, height=300, padx=2)
            executionFrame.pack(side=TOP, fill=BOTH)

            sbar = Scrollbar(self.execute_multiple_order_win,
                             orient="vertical", command=(canvas.yview))
            canvas.configure(yscrollcommand=sbar.set)
            sbar.pack(side=RIGHT, fill=Y)
            canvas.pack(side=LEFT, fill=BOTH)
            canvas.create_window((1,1), window=executionFrame, anchor="nw")
            executionFrame.bind(
                "<Configure>", lambda event, canvas=canvas: self.canvas_frame_reconfigurer(canvas))
        self.test_object = StringVar(value='LIMIT')
        self.form_object_consolidate = {}
        self.form_object_elements = {}
        self.form_object_price_elements = {}
        print(form_object_temp.keys(), 'all keys here')
        button_index = 0
        for key in form_object_temp.keys():
            # print(value,'\t\n')
            account_frame = LabelFrame(
                executionFrame, text=str(key), padx=10, pady=10)
            account_frame.pack(side=TOP, fill=X)
            price_frame = Frame(account_frame,padx=20)
            price_frame.pack(side=TOP,fill=X)
            Label(price_frame,text='Price').pack(side=LEFT,fill=NONE)
            self.form_object_price_elements[key] = Entry(price_frame)
            self.form_object_price_elements[key].pack(side=LEFT,fill=NONE)
            ref_var = IntVar()
            ref_var.set(0)
            self.form_object_price_elements[key].insert(0,form_object_price[key])
            limit_check = Checkbutton(price_frame,text='Limit',variable=ref_var)
            limit_check.pack(side=LEFT,fill=NONE)
            limit_check.configure(command=lambda : self.toggle_limit_form_keys(ref_var.get()))
            iceberg_var = IntVar()

            iceberg_check = Checkbutton(price_frame,text='Iceberg',variable=iceberg_var)
            iceberg_check.pack(side=LEFT,fill=NONE)
            iceberg_check.configure(
                command=lambda : self.toggle_multiple_iceberg_form_keys(iceberg_var.get())
                )
            for form_item in form_object_temp[key]:
                self.parent.update()
                button_index = button_index + 1
                print(form_item.values(),
                      'all keys here newest #######################\n')

                keychar = key + '#' + form_item['account'] + '#' + form_item['exit_producttype']
                self.form_object_consolidate[keychar] = {}
                self.form_object_elements[keychar] = []
                self.form_object_consolidate[keychar]['exit_ordertype'] = StringVar(
                )
                self.form_object_consolidate[keychar]['exit_producttype'] = StringVar(
                )
                self.form_object_consolidate[keychar]['exit_transactiontype'] = Variable(
                )
                self.form_object_consolidate[keychar]['quantity'] = StringVar()
                self.form_object_consolidate[keychar]['lots'] = StringVar()
                self.form_object_consolidate[keychar]['ltp'] = StringVar()
                self.form_object_consolidate[keychar]['instrument'] = form_item['instrument']
                self.form_object_consolidate[keychar]['account'] = form_item['account']
                self.form_object_consolidate[keychar]['total_quantity'] = IntVar()
                self.form_object_consolidate[keychar]['iceberg_legs'] = IntVar()
                self.form_object_consolidate[keychar]['iceberg_quantity'] = IntVar()

                self.form_object_consolidate[keychar]['exit_ordertype'].set(
                    form_item['exit_ordertype'])
                self.form_object_consolidate[keychar]['exit_producttype'].set(
                    form_item['exit_producttype'])
                self.form_object_consolidate[keychar]['exit_transactiontype'].set(
                    form_item['exit_transactiontype'])
                self.form_object_consolidate[keychar]['quantity'].set(
                    form_item['quantity'])
                self.form_object_consolidate[keychar]['lots'].set(
                    str(form_item['lots']))
                self.form_object_consolidate[keychar]['ltp'].set(
                    form_item['ltp'])
                self.form_object_consolidate[keychar]['total_quantity'].set(
                    form_item['total_quantity'])
                self.form_object_consolidate[keychar]['exchange'] = form_item['exchange']

                order_frame = Frame(account_frame, padx=10)
                order_frame.pack(side=TOP, fill=X)

                info_label = Label(order_frame, text=str(self.form_object_consolidate[keychar]['account']) + '\t' + self.form_object_consolidate[keychar]['exit_transactiontype'].get()
                                   + '\t', anchor='w')
                info_label.pack(side=TOP, fill=X)

                order_type_exit = LabelFrame(order_frame, text='Order type')
                rad1_exit = Radiobutton(order_type_exit, text='Market', var=(
                    self.form_object_consolidate[keychar]['exit_ordertype']
                ), value='MARKET')
                rad1_exit.pack(side=LEFT, fill=NONE)
                rad2_exit = Radiobutton(order_type_exit, text='Limit', variable=(
                    self.form_object_consolidate[keychar]['exit_ordertype']), value='LIMIT')
                rad2_exit.pack(side=LEFT, fill=NONE)
                order_type_exit.pack(side=LEFT, fill=NONE)

                self.form_object_elements[keychar] = []

                """ price_combo1 = Frame(order_frame,
                                        height=300, width=100, padx=1, pady=5)
                    labP = Label(price_combo1, width=10, text='Price')
                    self.form_object_elements[keychar].append(Entry(price_combo1, width=10, text='Enter Price',textvariable=self.form_object_consolidate[keychar]['ltp']))
                    labP.pack(side=LEFT)
                    self.form_object_elements[keychar][-1].pack(side=LEFT)
                    price_combo1.pack(side=LEFT, fill=NONE) """

                price_combo2 = Frame(order_frame,
                                     height=300, width=100, padx=1, pady=5)
                labQ = Label(price_combo2, width=10, text='Quantity')
                self.form_object_elements[keychar].append(Entry(
                    price_combo2, width=10, text='Enter Quantity', textvariable=self.form_object_consolidate[keychar]['quantity']))
                labQ.pack(side=LEFT)
                self.form_object_elements[keychar][-1].pack(side=LEFT)
                self.form_object_elements[keychar][-1].configure(
                    state='disable')
                # print(self.form_object_consolidate[keychar]['lots'].get(),'lot size \n')
                # print(button_index,'button_indexbutton_indexbutton_index',str(self.form_object_consolidate[keychar]['lots'].get()))
                labM = Label(price_combo2, width=7, text='Lot')
                self.form_object_elements[keychar].append(Spinbox(price_combo2, from_=0, to=50, values=(0, 10, 20, 30, 40, 50),
                                                                  width=5, wrap=False
                                                                  # ,command=(lambda keys=keychar: self.multiply_generic(keys))
                                                                  ))
                labM.pack(side=LEFT)

                self.form_object_elements[keychar][-1].pack(side=LEFT)
                self.form_object_elements[keychar][-1].config(state=NORMAL)
                self.form_object_elements[keychar][-1].delete(0, END)
                local_variable = self.form_object_consolidate[keychar]['lots'].get(
                )

                self.form_object_elements[keychar][-1].insert(
                    0, local_variable)
                print(self.form_object_elements[keychar][-1].get(), 'textetextextet >>>',
                      local_variable, 'x', self.form_object_consolidate[keychar]['lots'].get(), '\n')
              

                labIL = Label(price_combo2, width=5, text='Legs')
                self.form_object_elements[keychar].append(Entry(price_combo2, width=5, text='Enter Iceberg Legs',
                                                                textvariable=self.form_object_consolidate[keychar]['iceberg_legs']))
                labIL.pack(side=LEFT)
                self.form_object_elements[keychar][-1].pack(side=LEFT)
                # calculate_multiple_iceberg_lots_keys
                self.form_object_elements[keychar][-1].bind('<KeyRelease>', (
                    lambda e, keys=keychar, quant_ref=self.form_object_elements[keychar][-1]: self.calculate_multiple_iceberg_lots_keys(iceberg_var.get())))
                labIQ = Label(price_combo2, width=15, text='Iceberg Quantity')
                self.form_object_elements[keychar].append(Entry(price_combo2, width=5, text='Enter Iceberg Quantity',
                                                                textvariable=self.form_object_consolidate[keychar]['iceberg_quantity']))
                labIQ.pack(side=LEFT)
                self.form_object_elements[keychar][-1].pack(side=LEFT)

                labTQ = Label(price_combo2, width=15, text='Total Quantity')
                self.form_object_elements[keychar].append(Entry(price_combo2, width=10, text='Enter Total Quantity',
                                                                textvariable=self.form_object_consolidate[keychar]['total_quantity']))
                labTQ.pack(side=LEFT)
                self.form_object_elements[keychar][-1].pack(side=LEFT)

                price_combo2.pack(side=LEFT, fill=NONE)
                def combinedFuntion(self,keys,quant_ref,iceberg_var):
                    self.multiply_generic(keys, quant_ref) 
                    self.calculate_multiple_iceberg_lots_keys(iceberg_var.get())
                self.multiply_generic(
                    keychar, self.form_object_elements[keychar][-1])
                self.form_object_elements[keychar][1].configure(command=(
                    lambda keys=keychar, quant_ref=self.form_object_elements[keychar][-1]: combinedFuntion(self,keys,quant_ref,iceberg_var)))
                self.form_object_elements[keychar][1].bind('<KeyRelease>', (
                    lambda e, keys=keychar, quant_ref=self.form_object_elements[keychar][-1]: combinedFuntion(self,keys,quant_ref,iceberg_var)))
        execute_exit_button.config(command=self.run_multiple_exit)

    def toggle_limit_form_keys(self,event):
        for item in self.form_object_consolidate:
            self.form_object_consolidate[item]['exit_ordertype'].set('LIMIT' if event == 1 else 'MARKET')
 
    def toggle_multiple_iceberg_form_keys(self,event):
        print(self.form_object_elements,self.form_object_consolidate)
        for item in self.form_object_consolidate:
            if event == 1:
                quant = int(self.form_object_elements[item][1].get())
                if quant >= 5:
                    self.form_object_consolidate[item]['iceberg_legs'].set(3)
                    self.form_object_consolidate[item]['iceberg_quantity'].set(ceil((quant/3)) * int(self.form_object_consolidate[item]['quantity'].get()))
            else:
                self.form_object_consolidate[item]['iceberg_legs'].set(0)
                self.form_object_consolidate[item]['iceberg_quantity'].set(0)
    def calculate_multiple_iceberg_lots_keys(self,event):
        print(self.form_object_elements,self.form_object_consolidate)
        for item in self.form_object_consolidate:
            if event == 1:
                quant = int(self.form_object_elements[item][1].get())
                if quant >= 5:
                    self.form_object_consolidate[item]['iceberg_quantity'].set(
                        ceil((quant/self.form_object_consolidate[item]['iceberg_legs'].get()))
                         * int(self.form_object_consolidate[item]['quantity'].get()))
            else:
                self.form_object_consolidate[item]['iceberg_legs'].set(0)
                self.form_object_consolidate[item]['iceberg_quantity'].set(0)
    def run_multiple_exit(self):
        list_of_orders = []
        post_order_success = {}
        order_dump = {}
        sql_insertion_dump_mis = {}
        sql_insertion_dump_nrml = {}
        for item in self.form_object_elements.keys():
            account_id = item.split('#')[1]
            order_object = {
                'tradingsymbol': item.split('#')[0],
                'quantity': int(self.form_object_elements[item][-1].get()),
                'product': self.form_object_consolidate[item]['exit_producttype'].get(),
                'order_type': self.form_object_consolidate[item]['exit_ordertype'].get(),
                'transaction_type': self.form_object_consolidate[item]['exit_transactiontype'].get(),
                # 'price': float(self.form_object_consolidate[item]['ltp'].get()),
                'price': float(self.form_object_price_elements[item.split('#')[0]].get()),
                'validity': 'DAY',
                'variety': 'regular',
                'exchange': self.form_object_consolidate[item]['exchange']
            }
            variety_var = 'regular'
            if self.form_object_consolidate[item]['iceberg_quantity'].get() != '' and self.form_object_consolidate[item]['iceberg_quantity'].get() != None and int(self.form_object_consolidate[item]['iceberg_quantity'].get()) >= 5:
                order_object['variety'] = 'iceberg'
                variety_var = 'iceberg'
                order_object['iceberg_quantity'] = int(
                    self.form_object_consolidate[item]['iceberg_quantity'].get())
                order_object['iceberg_legs'] = int(
                    self.form_object_consolidate[item]['iceberg_legs'].get())
            # print(order_object,'>>>> \n')
            for acc in self.listOfAccounts:
                if acc.client_id == account_id:
                    
                    local_response = acc.exit_position(
                        order_object)
                    post_order_success[acc.client_id + '\t'+ order_object['product']] = copy.deepcopy(local_response)
                    # post_order_success[acc.client_id] = {'status': 'success', 'data': {'order_id': '230324202606793'}}
                    # if post_order_success[acc.client_id] != None and type(post_order_success[acc.client_id]) is dict and post_order_success[acc.client_id].get('status') and post_order_success[acc.client_id]['status'] == 'success':
                    #     self.single_order_exit_win.destroy()
                    #     pass
                    local_order_insertion_copy = copy.deepcopy(order_object)
                    # local_order_insertion_copy['variety'] = 'regular'
                    local_order_insertion_copy['transactiontype'] = local_order_insertion_copy['transaction_type']

                    if local_response != None and type(local_response) is dict and local_response.get('status') and local_response['status'] == 'success':
                        if order_object['product'] == 'NRML':
                            sql_insertion_dump_nrml[acc.client_id] = copy.deepcopy(
                                local_response)
                            sql_insertion_dump_nrml[acc.client_id]['exchange_order_status'] = 'OPEN PENDING'
                            sql_insertion_dump_nrml[acc.client_id]['order_quantity'] = int(self.form_object_elements[item][-1].get())
                            sql_insertion_dump_nrml[acc.client_id]['fill_quantity'] = 0
                            sql_insertion_dump_nrml[acc.client_id]['variety'] = variety_var
                        elif order_object['product'] == 'MIS':
                            sql_insertion_dump_mis[acc.client_id] = copy.deepcopy(
                                local_response)
                            sql_insertion_dump_mis[acc.client_id]['exchange_order_status'] = 'OPEN PENDING'
                            sql_insertion_dump_mis[acc.client_id]['order_quantity'] = int(self.form_object_elements[item][-1].get())
                            sql_insertion_dump_mis[acc.client_id]['fill_quantity'] = 0
                            sql_insertion_dump_mis[acc.client_id]['variety'] = variety_var
        local_order_insertion_date = datetime.now().strftime('%d/%m/%Y %H:%M:%S %f')
        print(json.dumps(sql_insertion_dump_nrml))
        if len(sql_insertion_dump_nrml.keys()) > 0:
            local_order_insertion_copy['product'] = 'NRML'
            self.order_db.insert_order((
                local_order_insertion_date,
                json.dumps(local_order_insertion_copy),
                json.dumps(sql_insertion_dump_nrml)
            ))
        local_order_insertion_date = datetime.now().strftime('%d/%m/%Y %H:%M:%S %f')
        
        if len(sql_insertion_dump_mis.keys()) > 0:
            local_order_insertion_copy['product'] = 'MIS'
            self.order_db.insert_order((
                local_order_insertion_date,
                json.dumps(local_order_insertion_copy),
                json.dumps(sql_insertion_dump_mis)
            ))
        self.loadPositionScreen()
        self.post_order_execeution_screen(post_order_success, 'Order Screen')
        self.execute_multiple_order_win.destroy()

    def multiply_generic(self, keychar, exit_total_quant):
        print(keychar, '<<<<<<<<keychar>>>>>>>')
        self.multiply_instrument_lots_generic(
            self.form_object_elements[keychar][0],
            exit_total_quant,
            self.form_object_elements[keychar][1],
            keychar
        )

    def execute_exit_position(self, account_id, order_object):
        account = ''
        response = ''
        print(order_object)
        variety_var = order_object['variety']
        post_order_success = {}
        for acc in self.listOfAccounts:
            if acc.client_id == account_id:
                post_order_success[acc.client_id] = acc.exit_position(
                    order_object)
                # post_order_success[acc.client_id] = {'status': 'success', 'data': {'order_id': '230324202606793'}}
                if post_order_success[acc.client_id] != None and type(post_order_success[acc.client_id]) is dict and post_order_success[acc.client_id].get('status') and post_order_success[acc.client_id]['status'] == 'success':
                    self.single_order_exit_win.destroy()
                    pass
        local_order_insertion_copy = copy.deepcopy(order_object)
        local_order_insertion_copy['variety'] = variety_var
        local_order_insertion_copy['transactiontype'] = local_order_insertion_copy['transaction_type']
        local_order_insertion_date = datetime.now().strftime('%d/%m/%Y %H:%M:%S %f')
        sql_insertion_dump = {}
        for key in post_order_success.keys():
            if post_order_success.get(key) and (type(post_order_success[key]) is dict) and post_order_success[key].get('status') and post_order_success[key]['status'] == 'success':
                sql_insertion_dump[key] = copy.deepcopy(
                    post_order_success[key])
                sql_insertion_dump[key]['exchange_order_status'] = 'OPEN PENDING'
                sql_insertion_dump[key]['order_quantity'] = int(local_order_insertion_copy['quantity'])
                sql_insertion_dump[key]['fill_quantity'] = 0
                sql_insertion_dump[key]['variety'] = variety_var
        if len(sql_insertion_dump.keys()) > 0:
            self.order_db.insert_order((
                local_order_insertion_date,
                json.dumps(local_order_insertion_copy),
                json.dumps(sql_insertion_dump)
            ))
            self.loadPositionScreen()
        self.post_order_execeution_screen(post_order_success, 'Order Screen')

        # self.recreate_running_orders_tree()
        # self.simulate_result({"type":"order","id":"","data":{"account_id":"ZL3443","unfilled_quantity":0,"checksum":"","placed_by":"ZL3443","order_id":"230324200948857","exchange_order_id":"2100000024359273","parent_order_id":'null',"status":"REJECTED","status_message":'null',"status_message_raw":'null',"order_timestamp":"2023-03-24 15:24:36","exchange_update_timestamp":"2023-03-24 15:24:36","exchange_timestamp":"2023-03-24 10:43:29","variety":"regular","exchange":"NFO","tradingsymbol":"TCS23APR3360CE","instrument_token":38196482,"order_type":"LIMIT","transaction_type":"BUY","validity":"DAY","product":"NRML","quantity":175,"disclosed_quantity":0,"price":6.25,"trigger_price":0,"average_price":0,"filled_quantity":0,"pending_quantity":175,"cancelled_quantity":175,"market_protection":0,"meta":{},"tag":'null',"guid":"66270XLptJ40MwQGlH"}})

    def recreate_open_position_tree_for_account(self, parent_tree_id):
        if hasattr(self, 'openPositionsTree') == False:
            return
        if len(self.openPositionsTree.get_children(parent_tree_id)) != 0:
            self.openPositionsTree.delete(
                *self.openPositionsTree.get_children(parent_tree_id))

        account = ''
        for acc in self.listOfAccounts:
            if acc.client_id == parent_tree_id and acc.authStatus == 'Logged In':
                account = acc
        positions = account.get_positions()
        print(type(positions) is dict)
        if account != '' and type(positions) is dict and positions.get('status') and positions['status'] == 'success':
            for position in positions['data']['net']:
                self.openPositionsTree.insert(account.client_id, 'end', iid=position['tradingsymbol'] + '#' + account.client_id + '#' + position['product'] + '#' +position['exchange'],
                                              tags=('content',),
                                              values=(
                    position['tradingsymbol'],
                    position['product'],
                    position['quantity'],
                    'Add',
                    'Exit'))

    def recreate_open_position_tree(self):
        if len(self.openPositionsTree.get_children()) != 0:
            self.openPositionsTree.delete(
                *self.openPositionsTree.get_children())

        pol = 1
        for account in self.listOfAccounts:
            if account.authStatus == 'Logged In':
                positions = account.get_positions()
                self.openPositionsTree.insert('', 'end', iid=account.client_id, text=pol,
                                            values=(
                                                account.client_id), open=True, tags=('order',))
                if (type(positions) is dict) and positions.get('status') and positions['status'] == 'success':
                    for position in positions['data']['net']:
                        self.openPositionsTree.insert(account.client_id, 'end', iid=position['tradingsymbol'] + '#' + account.client_id + '#' + position['product'] + '#' +position['exchange'],
                                                    tags=('content',),
                                                    values=(
                            position['tradingsymbol'],
                            position['product'],
                            position['quantity'],
                            'Add',
                            'Exit'))

            # pol += 1

    def find_exchange_order_id(self, order_id, account_id):
        exchange_order_id = ''
        list_of_orders = self.order_db.get_order_by_timestamp(order_id)
        if len(list_of_orders) > 0:
            loaded_order_json = json.loads(
                list_of_orders[0][2])[account_id]
            exchange_order_id = loaded_order_json['data']['order_id']
        return exchange_order_id

    def delete_order_single(self, exchange_order_id, account_number, order_id):
        if showwarning('Copy Trader', 'Are you sure you want to delete this order?', parent=self.view_order_win) == 'ok':
            print('execute delete')
            is_account_logged_in = False

            for acc in self.listOfAccounts:
                # print(acc.client_id,account_number,'delete_order_single')
                if acc.client_id == account_number and acc.authStatus == 'Logged In':
                    print('delete_order_single', acc)
                    is_account_logged_in = True
                    try:
                        deletion_response = acc.cancel_order(exchange_order_id)
                        showinfo('Cancel Order', deletion_response,
                                 parent=self.view_order_win)
                        if deletion_response['status'] == 'success':
                            self.update_order_status(
                                order_id, account_number, 'CANCELLED','exchange_order_status')

                    except Exception as e:
                        print(e)
                        showerror(
                            'Copy Trader', 'Something went wrong please try again later', parent=self.view_order_win)

            if is_account_logged_in == False:
                showerror(
                    'Copy Trader', "The account isn't logged in for order deletion", parent=self.view_order_win)

    def update_order_status(self, order_id, account_id, status,object_key,recreate = True):
        # exchange_order_id = ''
        list_of_orders = self.order_db.get_order_by_timestamp(order_id)
        if len(list_of_orders) > 0:
            loaded_order_wise_json = json.loads(
                list_of_orders[0][2])
            loaded_order_json = loaded_order_wise_json[account_id]
            if object_key in loaded_order_json :
                loaded_order_json[object_key] = status
                print(status, loaded_order_json[object_key], '1149')
        self.order_db.update_order_status(
            json.dumps(loaded_order_wise_json), order_id)
        if recreate :
            self.recreate_running_orders_tree()
        # print(loaded_order_wise_json)

    def modify_order_single(self, exchange_order_id, account_number,time_stamp):
        if self.is_order_modification_win_initial_load == False:
            self.order_modification_win.destroy()
            # self.view_order_win.update()
        if self.is_order_modification_win_initial_load == True:
            self.is_order_modification_win_initial_load = False


        list_of_orders = self.order_db.get_order_by_timestamp(time_stamp)              
        if len(list_of_orders) > 0:
            loaded_order_json = json.loads(list_of_orders[0][2])
        self.order_modification_win = Toplevel(self, padx=10,
                                               pady=10)

        self.mod_ordertype = StringVar()
        self.mod_ordertype.set('LIMIT')
        order_type = Frame(self.order_modification_win)
        rad1 = Radiobutton(order_type, text='Market', command=(
            lambda: self.orderType()), variable=(self.mod_ordertype), value='MARKET')
        rad1.pack(side=LEFT, fill=NONE)
        rad2 = Radiobutton(order_type, text='Limit', command=(
            lambda: self.orderType()), variable=(self.mod_ordertype), value='LIMIT')
        rad2.pack(side=RIGHT, fill=NONE)
        rad3 = Radiobutton(order_type, text='Stoploss', command=(
            lambda: self.orderType()), variable=(self.mod_ordertype), value='SL')
        rad3.pack(side=RIGHT, fill=NONE)
        # rad3.configure(state=DISABLED)
        order_type.pack(side=TOP, fill=X)
        price_combo1 = Frame(self.order_modification_win,
                             height=300, width=100, padx=1, pady=5)
        labP = Label(price_combo1, width=10, text='Price')
        self.modified_price = Entry(price_combo1, width=10, text='Enter Price')
        order_spec = json.loads(list_of_orders[0][2])
        print(order_spec[account_number].get('trigger_price'))

        try:
            ltp = float(self.get_last_traded_price(json.loads(
                        list_of_orders[0][1])['tradingsymbol'], False))
            self.modified_price.delete(0,END)
            self.modified_price.insert(0,ltp)
        except Exception as e:
            print(e)
        labP.pack(side=LEFT)
        self.modified_price.pack(side=RIGHT)
        price_combo1.pack(side=TOP, fill=X)
        price_combo2 = Frame(self.order_modification_win,
                             height=300, width=100, padx=1, pady=5)
        labQ = Label(price_combo2, width=10, text='Quantity')
        self.modified_quant = Entry(
            price_combo2, width=10, text='Enter Quantity')
        labQ.pack(side=LEFT)
        self.modified_quant.pack(side=RIGHT)
        try:
            print(loaded_order_json,'test',exchange_order_id)
            self.modified_quant.delete(0,END)
            self.modified_quant.insert(0, int(
                loaded_order_json[account_number]['order_quantity']))
        except Exception as e:
            print(e)
        price_combo2.pack(side=TOP, fill=X)
        # TODO: disable quantity
        order_variety = 'regular'
        try:
                price_combo3 = Frame(self.order_modification_win,
                             height=300, width=100, padx=1, pady=5)
                labSP = Label(price_combo3, width=10, text='Stop loss')
                self.stoploss_price = Entry(price_combo3,width=10, text = 'Stop Loss')
                labSP.pack(side=LEFT)
                self.stoploss_price.pack(side=RIGHT)
                price_combo3.pack(side=TOP,fill=X)
                if 'trigger_price' in order_spec[account_number]:
                    self.stoploss_price.delete(0,END)
                    self.stoploss_price.insert(0,float(order_spec[account_number]['trigger_price']))
                    self.mod_ordertype.set('SL')
                else:
                    self.stoploss_price.delete(0,END)
                    self.stoploss_price.insert(0,float(0))

                if 'variety' in order_spec[account_number] and order_spec[account_number]['variety'] == 'iceberg':
                    self.modified_quant.configure(state='disabled')
                    order_variety = 'iceberg'
        except Exception as e:
            print(e)
        Button(self.order_modification_win, text='Modify Order', width=20, command=(lambda: self.modify_order(self.mod_ordertype.get(),
                                                                                                              self.modified_price.get(),
                                                                                                              self.stoploss_price.get(),
                                                                                                              quantity=self.modified_quant.get(),
                                                                                                              account_number=account_number,
                                                                                                              exchange_order_id=exchange_order_id,
                                                                                                              time_stamp=time_stamp,
                                                                                                              order_variety=order_variety
                                                                                                              ))).pack(side=TOP, fill=X)

    def modify_order(self, order_type, price, trigger_price,quantity, account_number, exchange_order_id,time_stamp,order_variety):
        order_object = {}

        print(price, quantity, '1149')
        if order_type != None:
            order_object['order_type'] = order_type
        if price != None and price != '' and float(price) != 0:
            order_object['price'] = float(price)
        if quantity != None and quantity != '' and float(quantity) != 0:
            order_object['quantity'] = int(quantity)
        if trigger_price != None and trigger_price != '' and float(trigger_price) != 0:
            order_object['trigger_price'] = float(trigger_price)
        if order_variety == 'iceberg':
            del order_object['quantity']
        print(order_object, 'order object')
        if showwarning('Copy Trader', 'Are you sure you want to modify this order?', parent=self.view_order_win) == 'ok':
            print('execute modify')
            is_account_logged_in = False
            for acc in self.listOfAccounts:
                # print(acc.client_id,account_number,'delete_order_single')
                if acc.client_id == account_number and acc.authStatus == 'Logged In':
                    print('modify_order', acc)
                    is_account_logged_in = True
                    try:
                        mod_response = acc.modify_order(
                            exchange_order_id, order_object, order_variety)
                        showinfo('Modify Order', mod_response,
                                 parent=self.view_order_win)
                        if mod_response != None and mod_response['status'] == 'success':
                            if 'quantity' in order_object:
                                self.update_order_status(
                                                time_stamp, account_number,int(order_object['quantity']) ,'order_quantity')     
                                            
                            if 'trigger_price' in order_object:
                                self.update_order_status(
                                                    time_stamp, account_number,float(order_object['trigger_price']) ,'trigger_price')

                            self.order_modification_win.destroy()
                    except Exception as e:
                        print(e)
                        showerror(
                            'Copy Trader', 'Something went wrong please try again later', parent=self.view_order_win)
            if is_account_logged_in == False:
                showerror(
                    'Copy Trader', "The account isn't logged in for order modifcation", parent=self.view_order_win)

    def selectInstrument(self, riskpanel, quantity_panel, risk_setting):
        try:
            index = self.listBoxOfInstruments.curselection()  # on list double-click
            label = self.listBoxOfInstruments.get(index)
            instrument = ''
            for itme in [k for k in self.instrumentsData if label == k[1]]:
                print(itme, '405')
                instrument = itme
            # print(self.instrumentsData[index[0]])
            self.selectedInstrument.configure(state="normal")
            self.selectedInstrument.delete(0, END)
            self.selectedInstrument.insert(0, label)
            self.selectedInstrument.configure(state="disable")
            self.entL.configure(state="normal")
            self.entL.delete(0, END)
            self.entL.insert(0, int(instrument[5]))
            self.entL.configure(state="disable")
            self.multiplyLots(riskpanel, quantity_panel, risk_setting)
            self.selectedInstrumentData = instrument
            print("Selected", label,
                  instrument, instrument[5])
            self.update_last_traded_price()

        except:
            self.selectedInstrumentData = ''
            print("An exception occurred")

    def update_last_traded_price(self):
        try:
            index = self.listBoxOfInstruments.curselection()  # on list double-click
            label = self.listBoxOfInstruments.get(index)
            instrument = ''
            for itme in [k for k in self.instrumentsData if label == k[1]]:
                print(itme, '405')
                instrument = itme
            if instrument != '':
                self.entP.delete(0, END)
                self.entP.insert(
                    0, float(self.get_last_traded_price(instrument)))
        except Exception as e:
            print(e, '605')
            showerror('Copy Trader', 'Please select an instrument to fetch LTP')
            self.entP.delete(0, END)
            self.entP.insert(0, float(0))

    def get_last_traded_price(self, instrument, is_symbol_token=True):

        if len(self.listOfAccounts) != 0:
            for acc in self.listOfAccounts:
                if acc.authStatus == 'Logged In':
                    return acc.last_traded_price(instrument, is_symbol_token)
            return 0
        else:
            return 0

    def execute_order(self, accounts_object, quantity_panel, risk_panel, window_pane):

        print(datetime.now())
        for e in accounts_object.values():
            print(e.get(), 'iteration')

        index = self.listBoxOfInstruments.curselection()
        if len(index) == 0:
            showwarning(
                'Copy Trader', 'Please select an instrument to proceed', parent=window_pane)
            return

        # if isinstance(self.entM.get(),str):
        #     showwarning('Copy Trader','Please provide a valid value for multiples',master=window_pane)
        #     return
        print(index, 'cursor selection 527')
        label = self.listBoxOfInstruments.get(index)
        instrument = ''
        for item in [k for k in self.instrumentsData if label == k[1]]:
            print(item, '405')
            instrument = item
        order_quantity_by_account = {}
        print('Order :\n  variety = {0}\n  transactiontype = {1}\n  ordertype = {2}\n  producttype = {3} \n  duration = {4} \n  exchange = {5} \n  price = {6} \n  quantity = {7} \n  selected Instrument = {8} \n  symboltoken = {9} \n  tradingsymbol = {10} \n'.format(
            self.variety.get(), self.transactiontype.get(), self.ordertype.get(), self.producttype.get(), self.duration.get(), self.exchange.get(), self.entP.get(), self.entM.get(), instrument[2], instrument[1], instrument[0]))
        print(self.entSL.get(),'entSL')
        
        if self.ordertype.get() == 'SL':
            print(self.entSL.get(),'entSL')
        orderObject = {
            'variety': self.variety.get(),
            'tradingsymbol': instrument[1],
            'symboltoken': instrument[0],
            'transactiontype': self.transactiontype.get(),
            'exchange': self.exchange.get(),
            'ordertype': self.ordertype.get(),
            'producttype': self.producttype.get(),
            'duration': self.duration.get(),
            'price': self.entP.get(),
            'quantity': self.entQ.get(),
            'triggerprice': self.entSL.get(),
            'iceberg_legs': self.entIL.get(),
            'iceberg_quantity': self.entIQ.get(),
        }
        
        if len(self.listOfAccounts) == 0:
            showinfo('Copy Trader', 'No accounts to execute orders')
            return

        post_order_success = {}
        is_any_account_logged_in = False
        for acc in self.listOfAccounts:
            if accounts_object.get(acc.client_id) != None and accounts_object.get(acc.client_id).get() == True:

                if acc.get_auth_status() == 'Logged In':
                    is_any_account_logged_in = True
                    if risk_panel[acc.client_id].get() == True:
                        order_object_copy = copy.deepcopy(orderObject)
                        order_object_copy['quantity'] = str(
                            int(round(quantity_panel[acc.client_id].get())) *
                            int(self.entL.get())
                        )
                        order_quantity_by_account[acc.client_id] = int(
                            int(round(quantity_panel[acc.client_id].get())) *
                            int(self.entL.get())
                        )

                        order_object_copy['iceberg_legs'] = int(self.account_iceberg_lot[acc.client_id].get())
                        order_object_copy['iceberg_quantity'] = int(self.account_iceberg_quant[acc.client_id].get())
                        # print(order_object_copy,'557',quantity_panel[acc.client_id].get(),
                        # int(round(quantity_panel[acc.client_id].get())),int(self.entL.get()),
                        # int(round(quantity_panel[acc.client_id].get())) *
                        # int(self.entL.get())
                        # )
                        post_order_success[acc.client_id] = acc.place_order(
                            order_object_copy)
                        # post_order_success[acc.client_id] = {
                        #     "status": "success",
                        #     "data": {
                        #     "order_id": "151220000000000"
                        #     }
                        # }
                    else:
                        # print(orderObject,'561')
                        post_order_success[acc.client_id] = acc.place_order(
                            orderObject)
                        order_quantity_by_account[acc.client_id] = int(orderObject['quantity'])
                else:
                    # post_order_success[acc.client_id] = 'Not logged in to place orders'
                    pass
            else:
                # post_order_success[acc.client_id] = 'Inactive account'
                pass

        if is_any_account_logged_in == False:
            showinfo('Copy Trader', 'No accounts logged in to execute orders')
            return

        local_order_insertion_copy = copy.deepcopy(orderObject)
        local_order_insertion_date = datetime.now().strftime('%d/%m/%Y %H:%M:%S %f')
        # order_id TEXT PRIMARY KEY,
        # order_json TEXT NOT 'null',
        # order_collection TEXT NOT NULL
        sql_insertion_dump = {}
        for key in post_order_success.keys():
            if post_order_success.get(key) and (type(post_order_success[key]) is dict) and post_order_success[key].get('status') and post_order_success[key]['status'] == 'success':
                sql_insertion_dump[key] = copy.deepcopy(
                    post_order_success[key])
                sql_insertion_dump[key]['exchange_order_status'] = 'OPEN PENDING'
                sql_insertion_dump[key]['order_quantity'] = int(order_quantity_by_account[key])
                sql_insertion_dump[key]['fill_quantity'] = 0
                sql_insertion_dump[key]['variety'] = 'regular' if local_order_insertion_copy['variety'] == 'NORMAL' else 'iceberg'
                if local_order_insertion_copy['ordertype'] == 'SL':
                    sql_insertion_dump[key]['trigger_price'] = float(local_order_insertion_copy['triggerprice'])
        if len(sql_insertion_dump.keys()) > 0:
            self.order_db.insert_order((
                local_order_insertion_date,
                json.dumps(local_order_insertion_copy),
                json.dumps(sql_insertion_dump)
            ))
            self.loadPositionScreen()
        self.post_order_execeution_screen(post_order_success, 'Order Screen')
        self.recreate_running_orders_tree()

    def post_order_execeution_screen(self, accounts_order_object, title_action):
        win = Toplevel(self, height=500, width=700, padx=10,
                       pady=10)
        win.pack_propagate(0)
        print(title_action, '482')
        win.title(title_action)
        post_execution_tree = ttk.Treeview(win, column=('ITEM', 'VALUE'), show='headings',
                                           selectmode='extended', height=460)
        post_execution_tree.heading('ITEM', text='Item')
        post_execution_tree.heading('VALUE', text='Value')
        pol = 1
        for key, value in accounts_order_object.items():
            post_execution_tree.insert('', 'end', text=pol,
                                       values=(
                                           key, value), tags=('order',))
        post_execution_tree.pack(side=TOP, fill=BOTH)

    def searchInstruments(self):
        # print(self.searchInstrumentItem.get())
        # print([k for k in self.instrumentsData if self.searchInstrumentItem.get().lower() in k['symbol'].lower()])
        pos = 0
        self.listBoxOfInstruments.delete(0, END)
        for item in [k for k in self.instrumentsData if self.searchInstrumentItem.get().lower() in k[1].lower()]:
            self.listBoxOfInstruments.insert(
                pos, item[1])  # or insert(END,label)
            pos = pos + 1

    def orderType(self):
        print('CHecked')

   

    def onQuit(self):
        self.update()
        if askyesno('Copy Trader', 'Are you sure you want to exit?'):
            for item in self.listOfAccounts:
                if item.is_valid():
                    item.logout()
            self.quit()

    def fetchUserProfile():
        print('')

    def loadAccounts(self):
        name = askopenfilename(initialdir='.', filetypes=sheetTypes)
        # print(name)
        # name = '/home/jayant/Desktop/Accounts.xlsx'
        print(name, '472')
        loaded_accounts_object = {}
        if name:
            try:
                wb = load_workbook(name)
                ws = wb[wb.sheetnames[0]]
                for idx, row in enumerate(ws.rows):
                    if idx != 0:
                        values = []
                        for cell in row:
                            values.append(cell.value)
                        self.listOfXLSXAccounts.append(values)
                        temp_modded_account = list(values)
                        # acc_values[-1] = json.loads(acc_values[-1])
                        temp_modded_account[-1] = json.loads(temp_modded_account[-1])
                        acc = Account(*temp_modded_account)

                        is_duplicate = False
                        for x in self.listOfAccounts:
                            if x.client_id == acc.client_id:
                                is_duplicate = True

                        if is_duplicate == False and acc.is_valid():
                            acc.login()
                            print(acc.authStatus, acc.authStatus == 'Logged in')
                            if acc.authStatus == 'Logged In':
                                if self.account_db.insert_account_single_entry(
                                        acc.tuple_val()):
                                    self.load_accounts_from_db()
                                    
                                loaded_accounts_object[acc.client_id] = 'Loaded'
                            else:
                                loaded_accounts_object[acc.client_id] = 'Failed to load, please check the credentials and retry later'
                                pass
                self.post_order_execeution_screen(
                    loaded_accounts_object, 'Accounts')
                self.recreate_tree()
            except Exception as e:
                print('Exception',e)
                logging.warning('This will get logged to a file',e)
                if askyesno('Copy Trader', 'An error occured in the file, do you want to retry?'):
                    self.loadAccounts()
                else:
                    pass
        else:
            if askyesno('Copy Trader', 'No file selected, do you want to retry?'):
                self.loadAccounts()

    def load_accounts_from_db(self,initiate_serivce = False):
        if initiate_serivce:
            self.account_db = Account_DB()
        temp_accounts = self.account_db.get_accounts_data()
        print(temp_accounts, 'Temp accounts load')
        self.listOfAccounts = []
        if len(temp_accounts) != 0:
            for acc_values in temp_accounts:
                temp_modded_account = list(acc_values)
                # acc_values[-1] = json.loads(acc_values[-1])
                temp_modded_account[-1] = json.loads(temp_modded_account[-1])
                acc = Account(*temp_modded_account)
                if acc.is_valid():
                    self.listOfAccounts.append(acc)
                    # self.account_db.insert_account_single_entry(
                    #     acc.tuple_val())

        # self.initiate_login()

    def initiate_login(self):
        for item in self.listOfAccounts:
            print(item)
            if item.authStatus != 'Logged In':
                # item.login()
                item.get_user_profile()

    def initiateSocketConnections(self):
        for item in self.listOfAccounts:
            print(item)
            if item.is_valid():
                item.initiate_socket_connection()
            print("Accessing one single value (eg. CLIENT_ID): {0}".format(
                item.client_id))

    def createListOfAccountsWidget(self):
        self.tree = ttk.Treeview(self.canvas, column=('CLIENT_ID', 'STATUS', 'RISK','LOW','MEDIUM','HIGH'), show='headings',
                                 height=8,
                                 selectmode='browse',)
        self.tree.heading('CLIENT_ID', text='Client ID')
        self.tree.heading('STATUS', text='Status')
        self.tree.heading('RISK', text='Risk')
        self.tree.heading('LOW', text='Low')
        self.tree.heading('MEDIUM', text='Medium')
        self.tree.heading('HIGH', text='High')

        pos = 1
        col_width = self.tree.winfo_width()
        for acc in self.tree['columns']:
            self.tree.column(acc, anchor=CENTER, width=col_width)

        for acc in self.listOfAccounts:
            self.tree.insert('', 'end', text=pos, values=(acc.get_tree_view()))
            pos += 1

        treeScroll = ttk.Scrollbar(self.canvas)
        treeScroll.configure(command=(self.tree.yview))
        self.tree.configure(yscrollcommand=(treeScroll.set))
        self.tree.bind('<Button-3>', self.on_account_click)
        treeScroll.pack(side=RIGHT, fill=Y)
        self.tree.pack(side=LEFT, fill=BOTH, expand=True)

    def on_account_click(self, event):
        item = self.tree.identify('item', event.x, event.y)
        if self.tree.item(item) != '':
            if self.tree.item(item)['values']:
                selectedItem = self.tree.item(item)['values']
                # selected_account = [x for ind, x in enumerate(self.listOfAccounts) ]

                for ind, x in enumerate(self.listOfAccounts):
                    if x.client_id == selectedItem[0]:
                        selected_account = x
                print('you clicked on', selected_account)
                self.open_accounts_panel(selected_account, selectedItem)

    def open_accounts_panel(self, selected_account_object, selected_item_from_tree):
        win = Toplevel(self, padx=20, pady=20, width=300, height=300)
        win.pack_propagate(0)
        win.title('Account Configure')
        win.config()
        Label(win, text='{0}'.format(selected_account_object)).pack(
            side=TOP, fill=X, padx=10, pady=10)
        edit_delete_frame = Frame(win)
        edit_delete_frame.pack(side=TOP, fill=X, anchor=CENTER)
        Button(edit_delete_frame, text='Edit account',
               command=(lambda: ())).pack(side=LEFT)
        Button(edit_delete_frame, text='Delete account',
               command=(lambda: self.remove_account(
                   selected_account=selected_account_object, account_screen=win))
               ).pack(side=RIGHT)
        login_logout_frame = Frame(win)
        login_logout_frame.pack(side=TOP, fill=X, anchor=CENTER)

        login_button_instance = Button(login_logout_frame, text='Login')
        logout_button_instance = Button(login_logout_frame, text='Logut')
        login_button_instance.configure(command=(lambda:
                                                 self.single_account_login(selected_account_object, logout_button_instance)))
        logout_button_instance.configure(command=(lambda:
                                                  self.single_account_logout(selected_account_object, login_button_instance)))
        logout_button_instance.pack(side=RIGHT)
        login_button_instance.pack(side=LEFT)
        if selected_account_object.authStatus != 'Logged In':
            logout_button_instance.configure(state='disable')
        elif selected_account_object.authStatus == 'Logged In':
            login_button_instance.configure(state='disable')

    def single_account_login(self, account_object, logout_button):
        account_object.login()
        if account_object.authStatus == 'Logged In':
            account_object.start_thread(self.threaded_queue)
            self.recreate_tree()
            logout_button.configure(state='normal')
            self.place_order()
            self.loadPositionScreen()
            showinfo('Copy Trader', 'Logged in successfully')
        else:
            showerror('Copy Trader', 'Failed to login')

    def single_account_logout(self, account_object, login_button):
        account_object.logout()
        if account_object.authStatus == 'Logged Out':
            self.recreate_tree()
            login_button.configure(state='normal')
            self.place_order()
            self.loadPositionScreen()
            showinfo('Copy Trader', 'Logged out successfully')
        else:
            showerror('Copy Trader', 'Failed to logout')

    def remove_account(self, selected_account, account_screen):
        print(self.tree.get_children())
        if self.account_db.remove_account(selected_account.client_id) == True:
            if showinfo('Copy trader', 'Removed account {0} \n from copy trader'.format(selected_account),parent=account_screen) == 'ok':
                selected_index = -1
                for ind, obj in enumerate(self.listOfAccounts):
                    if obj.client_id == selected_account.client_id:
                        selected_index = ind
                if selected_index != -1:
                    self.listOfAccounts.pop(selected_index)
                account_screen.destroy()
                self.recreate_tree()
            else:
                showerror('Copy trader', 'Failed to remove {0} \n from copy trader'.format(
                    selected_account),parent=account_screen)
        pass

    def recreate_tree(self):
        if len(self.tree.get_children()) != 0:
            self.remove_all_treeitems()
        pos = 1
        for acc in self.listOfAccounts:
            self.tree.insert('', 'end', text=pos, values=(acc.get_tree_view()))
            pos += 1

    def remove_all_treeitems(self):
        self.tree.delete(*self.tree.get_children())

    def loadInstruments(self):
        # name = askopenfilename(initialdir='.', filetypes=jsonTypes)
        # print(name)
        self.instrumentsData = json.load(
            open('/home/jayant/CT/ScripMaster.json'))

    def loadAccountInfo(self):
        self.listOfAccounts

    def setup_database(self):
        self.call_progressbar()

        self.order_db = Orders()
        self.angel_db = AngelInstruments()
        self.zerodha_db = ZerodhaInstruments()
        last_date = self.angel_db.check_last_update_date()

        # print(last_date,'last date',datetime.strptime(last_date[0][1],'%d-%m-%Y'),datetime.today(),datetime.today().strftime('%d-%m-%Y'))
        print(last_date, 'angel last date')
        if (len(last_date) == 0) or (len(last_date) != 0 and len(last_date[0]) != 0 and datetime.strptime(last_date[0][0], '%d-%m-%Y') < datetime.strptime(datetime.today().strftime('%d-%m-%Y'), '%d-%m-%Y')):
            self.initiate_instrument_data_fetch_angel()

        last_date = self.zerodha_db.check_last_update_date()
        print(last_date, 'zerodha last date')
        if (len(last_date) == 0) or (len(last_date) != 0 and len(last_date[0]) != 0 and datetime.strptime(last_date[0][0], '%d-%m-%Y') < datetime.strptime(datetime.today().strftime('%d-%m-%Y'), '%d-%m-%Y')):
            self.initiate_instrument_data_fetch_zerodha()

        self.stop_progressbar()

    def initiate_instrument_data_fetch_zerodha(self):
        retries = 0
        success = False
        while not success and retries < 3:
            try:
                dump_connection = http.client.HTTPSConnection("api.kite.trade")
                dump_connection.request(
                    'GET', constant.INSTRUMENTCSVDUMPZERODHA)
                res = dump_connection.getresponse()
                data = res.read()
                if res.status == 200:
                    success = True
                    recordsData = []
                    # try:
                    data_list = csv.reader(data.decode(
                        'utf-8').splitlines(), delimiter='\n')
                    # except ValueError:
                    #         raise print("Couldn't parse the JSON response received from the server: {content}".format(
                    #             content=data))
                # sampleRecords = [{"token":"12487","symbol":"CYIENT-BL","name":"CYIENT","expiry":"","strike":"-1.000000","lotsize":"1","instrumenttype":"","exch_seg":"NSE","tick_size":"5.000000"},{"token":"972","symbol":"64KA30C-SG","name":"64KA30C","expiry":"","strike":"-1.000000","lotsize":"100","instrumenttype":"","exch_seg":"NSE","tick_size":"1.000000"},{"token":"12718","symbol":"RIIL-BL","name":"RIIL","expiry":"","strike":"-1.000000","lotsize":"1","instrumenttype":"","exch_seg":"NSE","tick_size":"5.000000"},{"token":"10959","symbol":"749HP34-SG","name":"749HP34","expiry":"","strike":"-1.000000","lotsize":"100","instrumenttype":"","exch_seg":"NSE","tick_size":"1.000000"}]
                    for elem in list(data_list):
                        # print(elem.values())
                        # print(tuple(elem[0].split(',')),'tuple lelemnt',elem)
                        recordsData.append(tuple(elem[0].split(',')))
                    recordsDataRowCount = self.zerodha_db.insert_instruments_data(
                        recordsData)
                    print(recordsDataRowCount)
                    res = self.zerodha_db.insert_date_entry()
                    print(res, 'check 1 zerodha db insertion')
                    self.stop_progressbar()
                    success = True
                    return

            except Exception as e:
                print("Error occured", e)
                retries += 1

    def initiate_instrument_data_fetch_angel(self):
        retries = 0
        success = False
        while not success and retries < 3:
            try:
                dump_connection = http.client.HTTPSConnection(
                    "margincalculator.angelbroking.com")
                dump_connection.request('GET', constant.INSTRUMENTCSVDUMP)
                res = dump_connection.getresponse()

                data = res.read()
                # print(res.status, data)
                if res.status == 200:
                    success = True
                    recordsData = []
                    try:
                        data_list = json.loads(data)
                    except ValueError:
                        raise print("Couldn't parse the JSON response received from the server: {content}".format(
                            content=data))
                # sampleRecords = [{"token":"12487","symbol":"CYIENT-BL","name":"CYIENT","expiry":"","strike":"-1.000000","lotsize":"1","instrumenttype":"","exch_seg":"NSE","tick_size":"5.000000"},{"token":"972","symbol":"64KA30C-SG","name":"64KA30C","expiry":"","strike":"-1.000000","lotsize":"100","instrumenttype":"","exch_seg":"NSE","tick_size":"1.000000"},{"token":"12718","symbol":"RIIL-BL","name":"RIIL","expiry":"","strike":"-1.000000","lotsize":"1","instrumenttype":"","exch_seg":"NSE","tick_size":"5.000000"},{"token":"10959","symbol":"749HP34-SG","name":"749HP34","expiry":"","strike":"-1.000000","lotsize":"100","instrumenttype":"","exch_seg":"NSE","tick_size":"1.000000"}]
                    for elem in data_list:
                        recordsData.append((
                            elem['token'],
                            elem['symbol'],
                            elem['name'],
                            elem['expiry'],
                            elem['strike'],
                            elem['lotsize'],
                            elem['instrumenttype'],
                            elem['exch_seg'],
                            elem['tick_size']
                        ))
                    recordsDataRowCount = self.angel_db.insert_instruments_data(
                        recordsData)
                    print(recordsDataRowCount, res.status)
                    res = self.angel_db.insert_date_entry()
                    print(res, 'check 1 angel db insertion')
                    self.stop_progressbar()
                    success = True
                    return

            except Exception as e:
                print("Error occured", e)
                retries += 1

    def call_progressbar(self):
        # self.prog_bar = ProgressBarPanel(self.parent)
        # self.prog_bar.start_progress()
        pass

    def stop_progressbar(self):
        # self.prog_bar.stop_progress()
        # self.prog_bar.remove_bar()
        pass

    def setup_default_risk_profiles(self):
        self.order_risk_profile_var = []
        self.account_risk_profile_var = []
        rows, cols = (3, 3)
        self.risk_db = RiskProfile()
        default_risk_profile = self.risk_db.get_risk_profile_by_name(
            'default_name')
        print(default_risk_profile)
        try:
            if len(default_risk_profile) != 0 and len(default_risk_profile[0]) != 0:
                print(json.loads(default_risk_profile[0][0]))
                loaded_profiles = json.loads(default_risk_profile[0][0])
                for i in range(rows):
                    self.order_risk_profile_var.append([])
                    for j in range(cols):
                        self.order_risk_profile_var[i].append(
                            DoubleVar(value=loaded_profiles['order_risk'][i][j]))

                for i in range(3):
                    self.account_risk_profile_var.append(
                        DoubleVar(value=loaded_profiles['account_risk'][i]))

            else:
                for i in range(rows):
                    self.order_risk_profile_var.append([])
                    for j in range(cols):
                        self.order_risk_profile_var[i].append(
                            DoubleVar(value=(i*20)+(j*18) + 20))

                self.account_risk_profile_var = []
                pos = 33.3
                for i in range(3):
                    self.account_risk_profile_var.append(DoubleVar(value=pos))
                    pos += 33.3
        except Exception as e:
            print(e, 'An Exception occured')

    def add_risk_rules(self):
        win = Toplevel(self, padx=20, height=500, width=350,
                       pady=20)
        win.pack_propagate(0)
        win.title('Account Risk Rules')
        win.config()
        # prev_panel = LabelFrame(win, text='Risk Profiles', pady=5, padx=10, width=250)
        # list_of_risk_profiles = self.risk_db.get_risk_profile_names()
        #         # setting variable for Integers
        # # for elem in list_of_risk_profiles:

        # selected_risk_profiles = StringVar()
        # selected_risk_profiles.set(list_of_risk_profiles[0])
        # # creating widget
        # account_risk_option = OptionMenu(
        #     prev_panel,
        #     selected_risk_profiles,
        #     *list_of_risk_profiles
        # )
        # # positioning widget
        # account_risk_option.pack(expand=True)

        # prev_panel.pack(side=TOP, fill=X)
        risk_matrix = LabelFrame(
            win, text='Account risk x Order risk matrix', pady=5, padx=10, width=250)
        Label(risk_matrix, text='Current profile',
              pady=20, padx=10, width=250).pack(side=TOP, fill=X)
        Label(risk_matrix, text='Low <------ Accounts ------> High',
              pady=20, padx=10, width=250).pack(side=TOP, fill=X)
        # text_var = []
        entries = []
        label_frames = []
        rows, cols = (3, 3)
        for i in range(rows):
            # append an empty list to your two arrays
            # so you can append to those later
            entries.append([])
            label_frames.append(LabelFrame(
                risk_matrix, text=self.return_risk_label(i), pady=5, padx=5))
            label_frames[i].pack(side=TOP)
            Label(label_frames[i], text='Low').pack(side=LEFT)
            for j in range(cols):
                # append your StringVar and Entry
                entries[i].append(
                    Entry(label_frames[i], textvariable=self.order_risk_profile_var[i][j], width=5))
                # print(i, j)
                entries[i][j].pack(side=LEFT, anchor=self.return_anchor(i, j))
            Label(label_frames[i], text='High').pack(side=LEFT)

        # button= Button(risk_matrix,text="Submit", bg='bisque3', width=15,
        # command=(lambda : self.save_risk_profile(text_var=text_var)))
        # button.place(x=160,y=140)
        risk_matrix.pack(side=TOP, fill=X)
        risk_account = LabelFrame(
            win, text='Independent account level risk settings', pady=5, padx=10, width=120)
        Label(risk_account, text='Low ------ Medium ------ High',
              pady=2, padx=10).pack(side=TOP, fill=X)

        for i in range(3):
            Entry(risk_account, textvariable=self.account_risk_profile_var[i], width=10).pack(
                side=LEFT, anchor=W)
        risk_account.pack(side=TOP, fill=X)

        button = Button(win, text="Save", width=15,
                        command=(lambda: self.save_risk_profile(text_var=self.order_risk_profile_var, risk_window=win)))
        button.pack(side=BOTTOM, fill=None)
        win.grab_set()

    def return_risk_label(self, i):
        if i == 0:
            return 'High Risk orders'
        elif i == 1:
            return 'Medium Risk orders'
        elif i == 2:
            return 'Low Risk orders'

    def return_anchor(self, i, j):
        anchor = ''
        if i == 0:
            anchor = anchor + 'n'
            if j == 0:
                anchor = anchor + 'w'
            if j == 2:
                anchor = anchor + 'e'
        if i == 1:
            anchor = anchor + 'center'
            if j == 0:
                anchor = 'w'
            if j == 2:
                anchor = 'e'
        if i == 2:
            anchor = anchor + 's'
            if j == 0:
                anchor = anchor + 'w'
            if j == 2:
                anchor = anchor + 'e'
        return anchor

    def save_risk_profile(self, text_var, risk_window):
        matrix = []
        rows, cols = (3, 3)
        for i in range(rows):
            matrix.append([])
            for j in range(cols):
                print(float(text_var[i][j].get()) == 0, float(text_var[i][j].get()) > 100, int(
                    text_var[i][j].get()) < 0, int(text_var[i][j].get()))
                if int(text_var[i][j].get()) > 100 or int(text_var[i][j].get()) < 0:
                    showwarning('Copy Trader',
                                'Please enter a value within 0 to 100')
                    return
                matrix[i].append(text_var[i][j].get())

        order_matrix = []
        for i in range(3):
            if int(self.account_risk_profile_var[i].get()) > 100 or int(self.account_risk_profile_var[i].get()) < 0:
                showwarning('Copy Trader',
                            'Please enter a value within 0 to 100')
                return
            order_matrix.append(self.account_risk_profile_var[i].get())

        risk_profile = {
            'order_risk': matrix,
            'account_risk': order_matrix
        }
        print(json.dumps(risk_profile))
        if self.risk_db.update_risk_profile_by_name('default_name', json.dumps(risk_profile)):
            if showinfo('Copy Trader', 'Risk profile updated successfully') == 'ok':
                risk_window.destroy()
        else:
            showerror('Copy Trader', 'Failed to update risk profile')

    def addAccountScreen(self):
        self.is_single_accountValid = False
        self.broker = StringVar()
        self.broker.set('angel')
        win = Toplevel(self, height=650, width=500, padx=20,
                       pady=20)
        win.pack_propagate(0)
        win.title('Add Account')
        win.config()
        Button(win, text='Import Accounts from a CSV',
               command=self.loadAccounts).pack(side=TOP, fill=X)
        Label(win, text='OR', fg='black', pady=15).pack(fill=X, side=TOP)
        main_account_input_frame = LabelFrame(win,
                                              text='Add a single account', pady=10, padx=10)
        main_account_input_frame.pack(side=TOP, fill=X)
        broker = LabelFrame(main_account_input_frame, text='Broker', pady=5)
        broker.pack(side=TOP, fill=X)
        radio1 = Radiobutton(broker, text='Angel', command=(
            lambda: ()), variable=self.broker, value='angel')
        radio1.pack(side=LEFT, fill=X)
        radio2 = Radiobutton(broker, text='Zerodha', command=(
            lambda: ()), variable=self.broker, value='zerodha')
        radio2.pack(side=LEFT, fill=X)
        clinet_id = LabelFrame(main_account_input_frame, text='Client ID',
                               pady=5,
                               padx=5)
        clinet_id.pack(side=TOP, anchor=W)
        self.clinet_id_entry = Entry(clinet_id,
                                     text='Enter client ID')
        self.clinet_id_entry.pack(side=TOP, fill=X)
        password = LabelFrame(main_account_input_frame, text='Password',
                              pady=5,
                              padx=5)
        password.pack(side=TOP, anchor=W)
        self.password_entry = Entry(password,
                                    text='Enter password')
        self.password_entry.pack(side=TOP, fill=X)
        api_key = LabelFrame(main_account_input_frame, text='API key',
                             pady=5,
                             padx=5)
        api_key.pack(side=TOP, anchor=W)
        self.api_key_entry = Entry(api_key,
                                   text='Enter API key')
        self.api_key_entry.pack(side=TOP, fill=X)
        secret_key = LabelFrame(main_account_input_frame, text='Secret key',
                                pady=5,
                                padx=5)
        secret_key.pack(side=TOP, anchor=W)
        self.secret_key_entry = Entry(secret_key,
                                      text='Enter secret key')
        self.secret_key_entry.pack(side=TOP, fill=X)
        totp_key = LabelFrame(main_account_input_frame, text='TOTP key',
                              pady=5,
                              padx=5)
        totp_key.pack(side=TOP, anchor=W)
        self.totp_key_entry = Entry(totp_key,
                                    text='Enter TOTP key')
        self.totp_key_entry.pack(side=TOP, fill=X)

        account_level_risk = LabelFrame(main_account_input_frame, text='Account level risk',
                                        pady=5,
                                        padx=5)
        risk_across_panel = LabelFrame(main_account_input_frame, text='Order level risk',
                                        pady=5,
                                        padx=5)
        self.order_level_risks_entry = {
            'low' : DoubleVar(),
            'medium' : DoubleVar(),
            'high' : DoubleVar()
        }
        self.order_level_risks_entry['low'].set(80)
        self.order_level_risks_entry['medium'].set(60)
        self.order_level_risks_entry['high'].set(40)
        Label(risk_across_panel,text='Low').pack(side=LEFT,fill=NONE)
        Entry(risk_across_panel,textvariable=self.order_level_risks_entry['low'],width=10).pack(side=LEFT,fill=NONE)
        Label(risk_across_panel,text='Medium').pack(side=LEFT,fill=NONE)
        Entry(risk_across_panel,textvariable=self.order_level_risks_entry['medium'],width=10).pack(side=LEFT,fill=NONE)
        Label(risk_across_panel,text='High').pack(side=LEFT,fill=NONE)
        Entry(risk_across_panel,textvariable=self.order_level_risks_entry['high'],width=10).pack(side=LEFT,fill=NONE)


        self.order_level_risks_entry
        risk_across_panel.pack(side=TOP, anchor=W)
        # setting variable for Integers
        self.account_risk = StringVar()
        self.account_risk.set(self.account_risk_vars[0])
        # creating widget
        account_risk_option = OptionMenu(
            account_level_risk,
            self.account_risk,
            *self.account_risk_vars,
            command=self.display_selected
        )
        # positioning widget
        account_risk_option.pack(expand=True)
        account_level_risk.pack(side=TOP, anchor=W)
        single_acc_save_btn = Button(
            main_account_input_frame, text='Save Account', command=(lambda : self.save_single_account(win)))
        single_acc_save_btn.pack(side=RIGHT, anchor=SE)
        single_acc_save_btn.configure(state='disable')
        Button(main_account_input_frame, text='Test Account', command=(
            lambda: self.validate_single_account(single_acc_save_btn))).pack(side=RIGHT, anchor=SE)

    def display_selected(self, choice):
        choice = self.account_risk.get()
        print(choice)

    def validate_single_account(self, single_acc_btn):
        if self.broker.get() == '':
            showerror('Missing parameter', 'Please select a broker')
            return False
        if self.clinet_id_entry.get() == '':
            showerror('Missing parameter', 'Please Enter client ID')
            return False
        if self.password_entry.get() == '':
            showerror('Missing parameter', 'Please Enter the password')
            return False
        if self.totp_key_entry.get() == '':
            showerror('Missing parameter', 'Please Enter the TOTP key')
            return False
        if self.secret_key_entry.get() == '':
            showerror('Missing parameter', 'Please Enter the secret key')
            return False
        else:
            print('broker:{0}\nclient id:{1}\npassword:{2}\nsecret key:{3}\ntotp key:{4}'.format(self.broker.get(
            ), self.clinet_id_entry.get(), self.password_entry.get(), self.secret_key_entry.get(), self.totp_key_entry.get()))
            self.temp_account_object = Account(self.clinet_id_entry.get(), self.password_entry.get(
            ), self.api_key_entry.get(), self.secret_key_entry.get(), self.totp_key_entry.get(), self.broker.get(), self.account_risk.get(),
            {
            'low' : self.order_level_risks_entry['low'].get(),
            'medium' : self.order_level_risks_entry['medium'].get(),
            'high' : self.order_level_risks_entry['high'].get()
            })
            for x in self.listOfAccounts:
                if x.client_id == self.clinet_id_entry.get():
                    showerror('Account already exists',
                              'Please retry with a different client id')
                    return None

            # print(temp_account_object)
            self.temp_account_object.login()
            if self.temp_account_object.authStatus == 'Logged In':
                showinfo('Test Account Validated',
                         'You can now save this account')
                single_acc_btn.configure(state='normal')
            else:
                showerror('Test Account Validation Failed',
                          'Please recheck your credetials or try again later')

    def save_single_account(self,win):

        try:
            if self.temp_account_object.authStatus == 'Logged In':
                self.account_db.insert_account_single_entry(
                    self.temp_account_object.tuple_val())
                self.listOfAccounts.append(self.temp_account_object)
                self.recreate_tree()
                self.place_order()
                if showinfo('Copy Trader','Account saved successfully') == 'ok':
                    win.destroy()
        except Exception as e:
            print(e)
            showwarning('An error occured please try again or contact the admin',
                        e, parent=(self.master))

    def start_progress_bar(self):
        # self.progbar = ttk.Progressbar(self.parent, orient=HORIZONTAL, length=220, mode="indeterminate")
        # self.progbar.pack(pady=20)
        # self.progbar.start()
        pass

    def stop_progress_bar(self):
        # if self.progbar:
        #     self.progbar.stop()
        pass


if __name__ == '__main__':

    root = Tk()
    root.title('Copy Trader 1.0')
    root.iconname('CT')
    # Label(root, text="Copy Trader").pack()
    height = root.winfo_screenheight()
    width = root.winfo_screenwidth()

    CopyTraderGUI(root, bd=3, relief=SUNKEN)
    root.mainloop()
